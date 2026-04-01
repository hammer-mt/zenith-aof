"""
Vendor Spend Forecasting & Maximizer Update Generator
Consolidates multiple blocking charts into a Maximizer-ready Excel forecast.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from copy import copy

OUTPUT_PATH = "/Users/hammermt/Codes/zenith-aof/exercise-2-deep-dive/ideas/07-forecasting/output.xlsx"

# --- Color constants ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
GREEN_FONT = Font(name="Calibri", color="006100", size=11)
RED_FONT = Font(name="Calibri", color="9C0006", size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CURRENCY_FMT = '"$"#,##0'
PCT_FMT = '0.0%'


def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_subheader_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_data_row_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def apply_total_row_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER


def color_delta(ws, row, col, value):
    """Apply green/red formatting to delta cells."""
    cell = ws.cell(row=row, column=col)
    if value > 0:
        cell.font = GREEN_FONT
        cell.fill = GREEN_FILL
    elif value < 0:
        cell.font = RED_FONT
        cell.fill = RED_FILL


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 1: Maximizer Update - VitalFit
# ============================================================

def build_vitalfit_sheet(wb):
    ws = wb.active
    ws.title = "Maximizer Update - VitalFit"

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "MAXIMIZER UPDATE -- VitalFit Q2 2026 (Apr-Jun)"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:L2")
    sub_cell = ws["A2"]
    sub_cell.value = "Generated: 2026-04-01 | Sources: Chart A (Mar 28), Chart B (Mar 15), Chart C (Mar 25) | Maximizer baseline: Feb 28"
    sub_cell.font = Font(name="Calibri", italic=True, size=10, color="595959")

    # Column headers - Row 4
    headers = [
        "Vendor",
        "Apr (New)", "May (New)", "Jun (New)",
        "Apr (Current)", "May (Current)", "Jun (Current)",
        "Apr Delta", "May Delta", "Jun Delta",
        "Source Chart", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4, 12)

    # Data: vendor -> (apr_new, may_new, jun_new, apr_cur, may_cur, jun_cur, source, notes)
    data = [
        ("Meta", 25000, 28000, 25000, 22000, 25000, 22000, "Chart A (Core Digital)", "Spring push increase"),
        ("Google Search", 16000, 18000, 16000, 15000, 16000, 15000, "Chart A (Core Digital)", ""),
        ("Google Shopping", 9000, 10000, 9000, 8000, 9000, 8000, "Chart A (Core Digital)", ""),
        ("Google PMAX", 14000, 14000, 12000, 12000, 13000, 12000, "Chart A (Core Digital)", ""),
        ("TikTok", 32100, 20150, 0, 18000, 18000, 15000, "Chart B (TikTok & Social)", "Spark Ads ends Apr 30; renewal TBD. Chart is 2 weeks old."),
        ("Creator Network (Grin)", 6000, 6200, 6000, 0, 0, 0, "Chart B (TikTok & Social)", "NEW vendor for Q2"),
        ("Pinterest", 2400, 4650, 4500, 0, 0, 0, "Chart B (TikTok & Social)", "Test-and-learn -- may not proceed"),
        ("Reddit", 0, 3100, 3000, 0, 0, 0, "Chart B (TikTok & Social)", "Test-and-learn -- may not proceed"),
        ("SparkPoint Media (CTV)", 15000, 17500, 17500, 10000, 10000, 10000, "Chart C (CTV & Programmatic)", "Increase per strong Q1 brand lift"),
        ("ReachMax Digital", 10500, 10500, 9000, 8000, 8000, 8000, "Chart C (CTV & Programmatic)", "Shifting to native-heavy mix"),
        ("The Trade Desk", 3000, 6000, 6000, 0, 0, 0, "Chart C (CTV & Programmatic)", "NEW -- self-serve programmatic test"),
        ("Vistar Media (DOOH)", 0, 6000, 6000, 0, 0, 0, "Chart C (CTV & Programmatic)", "Pending final approval"),
    ]

    row = 5
    totals_new = [0, 0, 0]
    totals_cur = [0, 0, 0]
    totals_delta = [0, 0, 0]

    for vendor, apr_n, may_n, jun_n, apr_c, may_c, jun_c, source, notes in data:
        apr_d = apr_n - apr_c
        may_d = may_n - may_c
        jun_d = jun_n - jun_c

        ws.cell(row=row, column=1, value=vendor)
        ws.cell(row=row, column=2, value=apr_n).number_format = CURRENCY_FMT
        ws.cell(row=row, column=3, value=may_n).number_format = CURRENCY_FMT
        ws.cell(row=row, column=4, value=jun_n).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=apr_c).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=may_c).number_format = CURRENCY_FMT
        ws.cell(row=row, column=7, value=jun_c).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8, value=apr_d).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9, value=may_d).number_format = CURRENCY_FMT
        ws.cell(row=row, column=10, value=jun_d).number_format = CURRENCY_FMT
        ws.cell(row=row, column=11, value=source)
        ws.cell(row=row, column=12, value=notes)

        # Color deltas
        color_delta(ws, row, 8, apr_d)
        color_delta(ws, row, 9, may_d)
        color_delta(ws, row, 10, jun_d)

        # Highlight conditional/pending rows
        if "Test-and-learn" in notes or "Pending" in notes:
            for c in range(1, 13):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

        apply_data_row_style(ws, row, 12)
        # Re-apply delta colors after data row style (which resets font)
        color_delta(ws, row, 8, apr_d)
        color_delta(ws, row, 9, may_d)
        color_delta(ws, row, 10, jun_d)
        if "Test-and-learn" in notes or "Pending" in notes:
            for c in range(1, 13):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

        totals_new[0] += apr_n
        totals_new[1] += may_n
        totals_new[2] += jun_n
        totals_cur[0] += apr_c
        totals_cur[1] += may_c
        totals_cur[2] += jun_c
        totals_delta[0] += apr_d
        totals_delta[1] += may_d
        totals_delta[2] += jun_d

        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=totals_new[0]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3, value=totals_new[1]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=4, value=totals_new[2]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=5, value=totals_cur[0]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=6, value=totals_cur[1]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=7, value=totals_cur[2]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=8, value=totals_delta[0]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=9, value=totals_delta[1]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=10, value=totals_delta[2]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=11, value="")
    ws.cell(row=row, column=12, value="")
    apply_total_row_style(ws, row, 12)

    # Column widths
    set_column_widths(ws, [26, 14, 14, 14, 14, 14, 14, 14, 14, 14, 28, 50])

    # Freeze panes (freeze below header row)
    ws.freeze_panes = "A5"


# ============================================================
# SHEET 2: Maximizer Update - LuminaSkin
# ============================================================

def build_luminaskin_sheet(wb):
    ws = wb.create_sheet("Maximizer Update - LuminaSkin")

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "MAXIMIZER UPDATE -- LuminaSkin Q2 2026 (Apr-Jun)"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    sub_cell = ws["A2"]
    sub_cell.value = "Generated: 2026-04-01 | Source: Chart D (Mar 22) | Maximizer baseline: Feb 28"
    sub_cell.font = Font(name="Calibri", italic=True, size=10, color="595959")

    headers = [
        "Vendor",
        "Apr (New)", "May (New)", "Jun (New)",
        "Apr (Current)", "May (Current)", "Jun (Current)",
        "Apr Delta", "May Delta", "Jun Delta",
        "Source Chart", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4, 12)

    # LuminaSkin Maximizer current values:
    # Meta: 15000, 15000, 15000
    # Google Search - Brand: 5000, 5000, 5000
    # Google Search - Non-Brand: 0, 0, 0
    # TikTok: 0, 0, 0
    # Influencer (AspireIQ): 0, 0, 0
    # DV360: 8000, 8000, 8000

    # New values from Chart D:
    # Meta (Always-On): 14000, 14000, 14000
    # Meta (Summer Launch): 0, 8000, 15000
    # Total Meta: 14000, 22000, 29000
    # But we should show them separately for clarity per the task spec.
    # The task says show:
    #   Meta (Always-On): Apr $14K, May $14K, Jun $14K
    #   Meta (Summer Launch): Apr $0, May $8K, Jun $15K
    # Current Maximizer for Meta total: 15000, 15000, 15000
    # We'll split the current across the two rows -- Always-On gets current, Summer gets 0.

    data = [
        ("Meta (Always-On)", 14000, 14000, 14000, 15000, 15000, 15000, "Chart D (Q2 Always-On)", "Reduced from $15K to $14K -- drift from Maximizer"),
        ("Meta (Summer Launch)", 0, 8000, 15000, 0, 0, 0, "Chart D (Summer Launch)", "NEW campaign. May is partial month (starts May 15)."),
        ("Google Search - Brand", 5000, 5000, 5000, 5000, 5000, 5000, "Chart D (Q2 Always-On)", "No change"),
        ("Google Search - Non-Brand", 10000, 12000, 12000, 0, 0, 0, "Chart D (Q2 Always-On)", "Ramping -- was paused in Q1, now reactivated"),
        ("TikTok (Summer Launch)", 0, 0, 12000, 0, 0, 0, "Chart D (Summer Launch)", "Starting Jun. Budget under review."),
        ("Influencer (AspireIQ)", 0, 5000, 5000, 0, 0, 0, "Chart D (Summer Launch)", "Fixed monthly retainer"),
        ("DV360 (Programmatic Display)", 6000, 6000, 6000, 8000, 8000, 8000, "Chart D (Q2 Always-On)", "Reduced rate in Q2 vs Maximizer"),
    ]

    row = 5
    totals_new = [0, 0, 0]
    totals_cur = [0, 0, 0]
    totals_delta = [0, 0, 0]

    for vendor, apr_n, may_n, jun_n, apr_c, may_c, jun_c, source, notes in data:
        apr_d = apr_n - apr_c
        may_d = may_n - may_c
        jun_d = jun_n - jun_c

        ws.cell(row=row, column=1, value=vendor)
        ws.cell(row=row, column=2, value=apr_n).number_format = CURRENCY_FMT
        ws.cell(row=row, column=3, value=may_n).number_format = CURRENCY_FMT
        ws.cell(row=row, column=4, value=jun_n).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=apr_c).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=may_c).number_format = CURRENCY_FMT
        ws.cell(row=row, column=7, value=jun_c).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8, value=apr_d).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9, value=may_d).number_format = CURRENCY_FMT
        ws.cell(row=row, column=10, value=jun_d).number_format = CURRENCY_FMT
        ws.cell(row=row, column=11, value=source)
        ws.cell(row=row, column=12, value=notes)

        color_delta(ws, row, 8, apr_d)
        color_delta(ws, row, 9, may_d)
        color_delta(ws, row, 10, jun_d)

        apply_data_row_style(ws, row, 12)
        color_delta(ws, row, 8, apr_d)
        color_delta(ws, row, 9, may_d)
        color_delta(ws, row, 10, jun_d)

        totals_new[0] += apr_n
        totals_new[1] += may_n
        totals_new[2] += jun_n
        totals_cur[0] += apr_c
        totals_cur[1] += may_c
        totals_cur[2] += jun_c
        totals_delta[0] += apr_d
        totals_delta[1] += may_d
        totals_delta[2] += jun_d

        row += 1

    # Totals
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=totals_new[0]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3, value=totals_new[1]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=4, value=totals_new[2]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=5, value=totals_cur[0]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=6, value=totals_cur[1]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=7, value=totals_cur[2]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=8, value=totals_delta[0]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=9, value=totals_delta[1]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=10, value=totals_delta[2]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=11, value="")
    ws.cell(row=row, column=12, value="")
    apply_total_row_style(ws, row, 12)

    set_column_widths(ws, [30, 14, 14, 14, 14, 14, 14, 14, 14, 14, 28, 55])
    ws.freeze_panes = "A5"


# ============================================================
# SHEET 3: Change Log
# ============================================================

def build_change_log(wb):
    ws = wb.create_sheet("Change Log")

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "CHANGE LOG -- All Changes vs. Current Maximizer (Feb 28 baseline)"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Client", "Vendor", "Month",
        "Current Value", "New Value", "Delta ($)", "Delta (%)",
        "Source Chart", "Updated By"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    apply_header_style(ws, 3, 9)

    # Build change log entries: only rows where delta != 0
    changes = []

    # VitalFit changes
    vf_data = [
        ("VitalFit", "Meta", "Apr", 22000, 25000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Meta", "May", 25000, 28000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Meta", "Jun", 22000, 25000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google Search", "Apr", 15000, 16000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google Search", "May", 16000, 18000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google Search", "Jun", 15000, 16000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google Shopping", "Apr", 8000, 9000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google Shopping", "May", 9000, 10000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google Shopping", "Jun", 8000, 9000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google PMAX", "Apr", 12000, 14000, "Chart A (Core Digital)", "Jordan"),
        ("VitalFit", "Google PMAX", "May", 13000, 14000, "Chart A (Core Digital)", "Jordan"),
        # Google PMAX Jun: 12000 -> 12000, no change, skip
        ("VitalFit", "TikTok", "Apr", 18000, 32100, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "TikTok", "May", 18000, 20150, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "TikTok", "Jun", 15000, 0, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Creator Network (Grin)", "Apr", 0, 6000, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Creator Network (Grin)", "May", 0, 6200, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Creator Network (Grin)", "Jun", 0, 6000, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Pinterest", "Apr", 0, 2400, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Pinterest", "May", 0, 4650, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Pinterest", "Jun", 0, 4500, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Reddit", "May", 0, 3100, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "Reddit", "Jun", 0, 3000, "Chart B (TikTok & Social)", "Anika"),
        ("VitalFit", "SparkPoint Media (CTV)", "Apr", 10000, 15000, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "SparkPoint Media (CTV)", "May", 10000, 17500, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "SparkPoint Media (CTV)", "Jun", 10000, 17500, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "ReachMax Digital", "Apr", 8000, 10500, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "ReachMax Digital", "May", 8000, 10500, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "ReachMax Digital", "Jun", 8000, 9000, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "The Trade Desk", "Apr", 0, 3000, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "The Trade Desk", "May", 0, 6000, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "The Trade Desk", "Jun", 0, 6000, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "Vistar Media (DOOH)", "May", 0, 6000, "Chart C (CTV & Programmatic)", "Marcus"),
        ("VitalFit", "Vistar Media (DOOH)", "Jun", 0, 6000, "Chart C (CTV & Programmatic)", "Marcus"),
    ]

    # LuminaSkin changes
    ls_data = [
        ("LuminaSkin", "Meta (Always-On)", "Apr", 15000, 14000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "Meta (Always-On)", "May", 15000, 14000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "Meta (Always-On)", "Jun", 15000, 14000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "Meta (Summer Launch)", "May", 0, 8000, "Chart D (Summer Launch)", "Taylor"),
        ("LuminaSkin", "Meta (Summer Launch)", "Jun", 0, 15000, "Chart D (Summer Launch)", "Taylor"),
        ("LuminaSkin", "Google Search - Non-Brand", "Apr", 0, 10000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "Google Search - Non-Brand", "May", 0, 12000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "Google Search - Non-Brand", "Jun", 0, 12000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "TikTok (Summer Launch)", "Jun", 0, 12000, "Chart D (Summer Launch)", "Taylor"),
        ("LuminaSkin", "Influencer (AspireIQ)", "May", 0, 5000, "Chart D (Summer Launch)", "Taylor"),
        ("LuminaSkin", "Influencer (AspireIQ)", "Jun", 0, 5000, "Chart D (Summer Launch)", "Taylor"),
        ("LuminaSkin", "DV360 (Programmatic Display)", "Apr", 8000, 6000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "DV360 (Programmatic Display)", "May", 8000, 6000, "Chart D (Q2 Always-On)", "Taylor"),
        ("LuminaSkin", "DV360 (Programmatic Display)", "Jun", 8000, 6000, "Chart D (Q2 Always-On)", "Taylor"),
    ]

    all_changes = vf_data + ls_data

    row = 4
    for client, vendor, month, cur_val, new_val, source, updated_by in all_changes:
        delta = new_val - cur_val
        if delta == 0:
            continue
        delta_pct = delta / cur_val if cur_val != 0 else None

        ws.cell(row=row, column=1, value=client)
        ws.cell(row=row, column=2, value=vendor)
        ws.cell(row=row, column=3, value=month)
        ws.cell(row=row, column=4, value=cur_val).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=new_val).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=delta).number_format = CURRENCY_FMT
        if delta_pct is not None:
            ws.cell(row=row, column=7, value=delta_pct).number_format = PCT_FMT
        else:
            ws.cell(row=row, column=7, value="NEW")
        ws.cell(row=row, column=8, value=source)
        ws.cell(row=row, column=9, value=updated_by)

        apply_data_row_style(ws, row, 9)
        color_delta(ws, row, 6, delta)

        row += 1

    set_column_widths(ws, [16, 32, 10, 16, 16, 16, 12, 30, 14])
    ws.freeze_panes = "A4"


# ============================================================
# SHEET 4: Anomaly Flags
# ============================================================

def build_anomaly_flags(wb):
    ws = wb.create_sheet("Anomaly Flags")

    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "ANOMALY FLAGS -- Items Requiring Investigation Before Maximizer Entry"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["#", "Category", "Client", "Detail", "Recommended Action"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    apply_header_style(ws, 3, 5)

    anomalies = [
        (1, "Stale Source Data", "VitalFit",
         "Blocking Chart B (TikTok & Social) was last updated Mar 15 -- 17 days ago as of Apr 1. TikTok spend numbers may have changed since then.",
         "Contact Anika (Insights) for updated Chart B before entering TikTok numbers."),
        (2, "Flight Expiration -- Renewal Unknown", "VitalFit",
         "TikTok Spark Ads (VF Spring Spark Ads) flight ends Apr 30. The blocking chart does not indicate whether this campaign renews for May/Jun.",
         "Confirm with Anika whether Spark Ads renews or stops. Current forecast shows $0 for May/Jun."),
        (3, "Conditional Spend -- Pinterest", "VitalFit",
         "Pinterest test-and-learn campaign ($2,400 Apr, $4,650 May, $4,500 Jun) is conditional on Q1 performance review results. 'May not proceed.'",
         "Flag as conditional in Maximizer. Check with Anika for Q1 review outcome before confirming entry."),
        (4, "Conditional Spend -- Reddit", "VitalFit",
         "Reddit community campaign ($3,100 May, $3,000 Jun) is conditional on Q1 performance review. 'May not proceed.'",
         "Same as Pinterest -- hold entry pending Q1 review confirmation."),
        (5, "Pending Approval -- Vistar DOOH", "VitalFit",
         "Vistar Media DOOH ($6,000 May, $6,000 Jun) is described as 'pending final approval -- gym network OOH.' Not yet confirmed spend.",
         "Enter as $0 until approval is confirmed. Track separately as pipeline item."),
        (6, "Rate Change Risk -- TikTok", "VitalFit",
         "Chart B notes that 'TikTok rates may increase in May per managed service renegotiation.' Current forecast uses the existing $650/day rate.",
         "If renegotiation happens, May/Jun TikTok numbers will need revision. Monitor."),
        (7, "Maximizer Drift -- Meta", "LuminaSkin",
         "LuminaSkin Meta shows $15,000/mo in Maximizer but blocking chart (Chart D) shows $14,000/mo for Always-On. $1,000/mo discrepancy.",
         "Confirm with Taylor whether the rate truly dropped. Update Maximizer to $14K if confirmed."),
        (8, "Maximizer Drift -- DV360", "LuminaSkin",
         "LuminaSkin DV360 shows $8,000/mo in Maximizer but Chart D shows $6,000/mo ('renewed at reduced rate in Q2'). $2,000/mo discrepancy.",
         "Confirm reduced rate with Taylor. Update Maximizer from $8K to $6K."),
        (9, "Reactivated Spend -- NB Search", "LuminaSkin",
         "LuminaSkin Google Search Non-Brand was $0 in Maximizer (was paused) but now ramping to $10K-12K/mo per Chart D. Significant new spend.",
         "Verify reactivation with Taylor. Enter $10K Apr, $12K May, $12K Jun per chart."),
    ]

    row = 4
    for num, cat, client, detail, action in anomalies:
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=cat)
        ws.cell(row=row, column=3, value=client)
        ws.cell(row=row, column=4, value=detail)
        ws.cell(row=row, column=5, value=action)

        apply_data_row_style(ws, row, 5)

        # Color-code by category
        if "Stale" in cat or "Rate Change" in cat:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = ORANGE_FILL
        elif "Conditional" in cat or "Pending" in cat:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = YELLOW_FILL
        elif "Drift" in cat or "Reactivated" in cat:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = LIGHT_RED_FILL
        elif "Flight" in cat:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

        # Wrap text for detail and action columns
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 50

        row += 1

    set_column_widths(ws, [5, 30, 14, 65, 55])
    ws.freeze_panes = "A4"


# ============================================================
# SHEET 5: Stale Data Warnings
# ============================================================

def build_stale_data_warnings(wb):
    ws = wb.create_sheet("Stale Data Warnings")

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "STALE DATA WARNINGS -- Source Chart Freshness Assessment"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Chart Name", "Last Updated", "Days Since Update", "Prepared By", "Risk Level", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    apply_header_style(ws, 3, 6)

    today = date(2026, 4, 1)

    charts = [
        ("Chart A: VitalFit Core Digital", date(2026, 3, 28), "Jordan (Planner)", "LOW",
         "Updated 4 days ago. Current and reliable."),
        ("Chart B: VitalFit TikTok & Social", date(2026, 3, 15), "Anika (Insights)", "HIGH",
         "Updated 17 days ago. TikTok rates may have changed. Request updated chart immediately."),
        ("Chart C: VitalFit CTV & Programmatic", date(2026, 3, 25), "Marcus (Programmatic Lead)", "LOW",
         "Updated 7 days ago. Recent enough for forecasting."),
        ("Chart D: LuminaSkin Q2", date(2026, 3, 22), "Taylor (Planner)", "MEDIUM",
         "Updated 10 days ago. Acceptable but approaching staleness threshold."),
        ("Maximizer Current Entries", date(2026, 2, 28), "Multiple", "BASELINE",
         "Feb 28 baseline. This is the comparison point, not a source chart."),
    ]

    row = 4
    for chart_name, last_updated, prepared_by, risk, notes in charts:
        days_since = (today - last_updated).days

        ws.cell(row=row, column=1, value=chart_name)
        ws.cell(row=row, column=2, value=last_updated.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=3, value=days_since)
        ws.cell(row=row, column=4, value=prepared_by)
        ws.cell(row=row, column=5, value=risk)
        ws.cell(row=row, column=6, value=notes)

        apply_data_row_style(ws, row, 6)

        # Color by risk level
        if risk == "HIGH":
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = LIGHT_RED_FILL
        elif risk == "MEDIUM":
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = YELLOW_FILL
        elif risk == "LOW":
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = GREEN_FILL

        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 35

        row += 1

    # Legend
    row += 1
    ws.cell(row=row, column=1, value="Risk Level Thresholds:").font = BOLD_FONT
    row += 1
    ws.cell(row=row, column=1, value="LOW: 0-7 days since update").font = NORMAL_FONT
    ws.cell(row=row, column=1).fill = GREEN_FILL
    row += 1
    ws.cell(row=row, column=1, value="MEDIUM: 8-14 days since update").font = NORMAL_FONT
    ws.cell(row=row, column=1).fill = YELLOW_FILL
    row += 1
    ws.cell(row=row, column=1, value="HIGH: 15+ days since update -- request refresh before entering data").font = NORMAL_FONT
    ws.cell(row=row, column=1).fill = LIGHT_RED_FILL

    set_column_widths(ws, [38, 16, 20, 28, 14, 60])
    ws.freeze_panes = "A4"


# ============================================================
# MAIN
# ============================================================

def main():
    wb = openpyxl.Workbook()
    build_vitalfit_sheet(wb)
    build_luminaskin_sheet(wb)
    build_change_log(wb)
    build_anomaly_flags(wb)
    build_stale_data_warnings(wb)
    wb.save(OUTPUT_PATH)
    print(f"Output saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
