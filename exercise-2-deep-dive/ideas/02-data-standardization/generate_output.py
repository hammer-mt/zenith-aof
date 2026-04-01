"""
Generate MDI Standardized Data Excel file from raw vendor exports.
Reads Meta, Google, and TikTok data, transforms to MDI format, and outputs
a professionally formatted Excel workbook.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import csv
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_PATH = os.path.join(BASE_DIR, "output.xlsx")

# ============================================================
# COLOR AND STYLE DEFINITIONS
# ============================================================
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

DATE_RANGE = "2026-03-01 to 2026-03-15"


def read_csv(filename):
    path = os.path.join(DATA_DIR, filename)
    with open(path, "r") as f:
        reader = csv.DictReader(f)
        return list(reader)


def safe_float(val):
    if val is None:
        return 0.0
    return float(str(val).replace(",", "").replace("$", "").strip())


def safe_int(val):
    if val is None:
        return 0
    return int(float(str(val).replace(",", "").strip()))


# ============================================================
# TRANSFORM META DATA
# ============================================================
def transform_meta(rows):
    """
    Meta: Amount Spent = Spend_Media (does NOT include fees).
    Agency fee = 15% of media spend.
    Spend_Total = Spend_Media + Spend_Fees.
    Link Clicks = Clicks.
    Results = Conversions.
    """
    results = []
    mapping_decisions = []
    needs_review = []

    # Ad format mapping
    format_map = {
        "Single Image": "Standard_Display",
        "Carousel": "Standard_Display",
        "Video": "Video",
        "Dynamic Product": "Native",
    }

    # Conv type mapping
    conv_type_map = {
        "Purchase": "Purchase",
        "Add to Cart": "AddToCart",
    }

    # Campaign name mapping based on original campaign names
    campaign_name_map = {
        ("VF_Spring_Prosp_Broad_18-35", "Fitness Enthusiasts"): "VitalFit_Meta_Prospecting_FitnessEnthusiasts_Q1-2026",
        ("VF_Spring_Prosp_Broad_18-35", "Health Conscious"): "VitalFit_Meta_Prospecting_HealthConscious_Q1-2026",
        ("VF_Spring_Prosp_LAL", "LAL - Past Purchasers"): "VitalFit_Meta_Prospecting_LAL-PastPurchasers_Q1-2026",
        ("VF_Spring_Retarget_ATC", "Site Visitors - ATC"): "VitalFit_Meta_Retargeting_SiteVisitors-ATC_Q1-2026",
    }

    # Record mapping decisions
    mapping_decisions.append(("Meta", "Amount Spent", "Spend_Media", "Direct mapping. Meta Amount Spent excludes agency fees per vendor reference."))
    mapping_decisions.append(("Meta", "Agency Fee (15%)", "Spend_Fees", "Calculated as 15% of Spend_Media. Fee rate specified in export footnote."))
    mapping_decisions.append(("Meta", "Link Clicks", "Clicks", "Used Link Clicks (not All Clicks). All Clicks includes reactions/comments per vendor reference."))
    mapping_decisions.append(("Meta", "Results", "Conversions", "Direct mapping per vendor reference."))
    mapping_decisions.append(("Meta", "Result Type: Purchase", "Conv_Type: Purchase", "Direct mapping."))
    mapping_decisions.append(("Meta", "Result Type: Add to Cart", "Conv_Type: AddToCart", "Mapped to MDI standard 'AddToCart' (no spaces)."))
    mapping_decisions.append(("Meta", "Reach", "(not mapped)", "No MDI equivalent per vendor reference. Logged but excluded."))
    mapping_decisions.append(("Meta", "Frequency", "(not mapped)", "No MDI equivalent per vendor reference. Logged but excluded."))
    mapping_decisions.append(("Meta", "All Clicks", "(not mapped)", "Excluded in favor of Link Clicks per vendor reference."))
    mapping_decisions.append(("Meta", "Format: Single Image", "Ad_Product: Standard_Display", "Standard single-image ad mapped to Standard_Display."))
    mapping_decisions.append(("Meta", "Format: Carousel", "Ad_Product: Standard_Display", "Multi-image carousel mapped to Standard_Display (display format, not video)."))
    mapping_decisions.append(("Meta", "Format: Video", "Ad_Product: Video", "Direct mapping."))
    mapping_decisions.append(("Meta", "Format: Dynamic Product", "Ad_Product: Native", "Dynamic product ads mapped to Native (personalized/dynamic format)."))

    for row in rows:
        campaign_raw = row["Campaign name"].strip()
        ad_set = row["Ad Set"].strip()
        fmt = row["Format"].strip()
        impressions = safe_int(row["Impressions"])
        clicks = safe_int(row["Link Clicks"])
        spend_media = safe_float(row["Amount Spent"])
        conversions = safe_int(row["Results"])
        result_type = row["Result Type"].strip()

        spend_fees = round(spend_media * 0.15, 2)
        spend_total = round(spend_media + spend_fees, 2)

        # Campaign name standardization
        key = (campaign_raw, ad_set)
        if key in campaign_name_map:
            campaign_name = campaign_name_map[key]
        else:
            # For duplicate ad set names (retarget has two rows with same ad set),
            # use the same standardized name
            for k, v in campaign_name_map.items():
                if k[0] == campaign_raw:
                    campaign_name = v
                    break
            else:
                campaign_name = f"VitalFit_Meta_Unknown_{ad_set}_Q1-2026"
                needs_review.append(("Meta", campaign_raw, "Campaign_Name", f"Could not confidently map campaign '{campaign_raw}' / ad set '{ad_set}' to MDI naming convention."))

        ad_product = format_map.get(fmt)
        if ad_product is None:
            ad_product = "NEEDS_REVIEW"
            needs_review.append(("Meta", fmt, "Ad_Product", f"Unknown Meta format '{fmt}' has no MDI mapping."))

        conv_type = conv_type_map.get(result_type)
        if conv_type is None:
            conv_type = "NEEDS_REVIEW"
            needs_review.append(("Meta", result_type, "Conv_Type", f"Unknown conversion type '{result_type}'."))

        # Calculated metrics
        cpm = round((spend_media / impressions) * 1000, 2) if impressions > 0 else 0
        cpc = round(spend_media / clicks, 2) if clicks > 0 else 0
        cpa = round(spend_media / conversions, 2) if conversions > 0 else 0
        ctr = round((clicks / impressions) * 100, 2) if impressions > 0 else 0

        results.append({
            "Campaign_Name": campaign_name,
            "Platform": "Meta",
            "Ad_Product": ad_product,
            "Date_Range": DATE_RANGE,
            "Impressions": impressions,
            "Clicks": clicks,
            "Spend_Media": spend_media,
            "Spend_Fees": spend_fees,
            "Spend_Total": spend_total,
            "Conversions": conversions,
            "Conv_Type": conv_type,
            "CPM": cpm,
            "CPC": cpc,
            "CPA": cpa,
            "CTR": ctr,
        })

    return results, mapping_decisions, needs_review


# ============================================================
# TRANSFORM GOOGLE DATA
# ============================================================
def transform_google(rows):
    """
    Google: Cost = Spend_Total (INCLUDES fees).
    Agency fee = flat $2,500/month. Prorated for half month (Mar 1-15) = $1,250.
    Distribute $1,250 across campaigns proportional to spend share.
    Spend_Media = Cost - Spend_Fees.
    """
    results = []
    mapping_decisions = []
    needs_review = []

    # Calculate total cost for proration
    total_cost = sum(safe_float(r["Cost"]) for r in rows)
    monthly_fee = 2500.00
    # Mar 1-15 = 15 days out of 31 in March
    prorated_fee = round(monthly_fee * (15 / 31), 2)  # $1,209.68

    # Ad type mapping
    ad_type_map = {
        "Responsive Search": "Search",
        "Shopping": "Shopping",
    }

    # Conv type mapping
    conv_type_map = {
        "Purchase": "Purchase",
        "Lead": "Lead",
    }

    # Campaign name mapping
    campaign_name_map = {
        ("VF_Spring_Brand", "Brand - Exact"): "VitalFit_Google_Brand_BrandExact_Q1-2026",
        ("VF_Spring_Brand", "Brand - Broad"): "VitalFit_Google_Brand_BrandBroad_Q1-2026",
        ("VF_Spring_NB_Protein", "Non-Brand - Protein"): "VitalFit_Google_Acquisition_Protein_Q1-2026",
        ("VF_Spring_NB_Supplements", "Non-Brand - Supplements"): "VitalFit_Google_Acquisition_Supplements_Q1-2026",
        ("VF_Spring_Shopping", "All Products"): "VitalFit_Google_Shopping_AllProducts_Q1-2026",
        ("VF_Spring_PMAX", "Performance Max"): "VitalFit_Google_PMAX_Broad_Q1-2026",
    }

    # Record mapping decisions
    mapping_decisions.append(("Google", "Cost", "Spend_Total", "Google Cost includes fees per vendor reference. Used as Spend_Total basis."))
    mapping_decisions.append(("Google", "Agency Fee ($2,500/mo)", "Spend_Fees", f"Flat $2,500/month prorated to 15/31 days = ${prorated_fee:.2f}. Distributed across campaigns by spend share."))
    mapping_decisions.append(("Google", "Impr.", "Impressions", "Abbreviated column name mapped per vendor reference."))
    mapping_decisions.append(("Google", "Clicks", "Clicks", "Direct mapping per vendor reference."))
    mapping_decisions.append(("Google", "Conversions", "Conversions", "Direct mapping per vendor reference."))
    mapping_decisions.append(("Google", "Conv. value", "(not mapped)", "No MDI equivalent per vendor reference. Logged but excluded."))
    mapping_decisions.append(("Google", "Ad type: Responsive Search", "Ad_Product: Search", "Responsive Search Ads are the standard Google Search format."))
    mapping_decisions.append(("Google", "Ad type: Shopping", "Ad_Product: Shopping", "Direct mapping."))
    mapping_decisions.append(("Google", "Ad type: PMAX", "Ad_Product: Standard_Display", "Performance Max spans multiple placements. Mapped to Standard_Display as closest MDI category. Flagged for review."))
    mapping_decisions.append(("Google", "Conv. type: Purchase", "Conv_Type: Purchase", "Direct mapping."))

    # Flag PMAX mapping
    needs_review.append(("Google", "VF_Spring_PMAX", "Ad_Product", "Performance Max (PMAX) is a multi-format campaign type spanning Search, Display, YouTube, Discover, and Shopping. Mapped to 'Standard_Display' as default but this may underrepresent the campaign's reach. Consider creating a separate 'PMAX' Ad_Product category in the MDI schema."))

    for row in rows:
        campaign_raw = row["Campaign"].strip()
        ad_group = row["Ad group"].strip()
        ad_type = row["Ad type"].strip()
        impressions = safe_int(row["Impr."])
        clicks = safe_int(row["Clicks"])
        cost = safe_float(row["Cost"])
        conversions = safe_int(row["Conversions"])
        conv_type_raw = row["Conv. type"].strip()

        # Calculate fee allocation for this campaign
        spend_share = cost / total_cost if total_cost > 0 else 0
        spend_fees = round(prorated_fee * spend_share, 2)
        spend_total = cost
        spend_media = round(spend_total - spend_fees, 2)

        # Campaign name standardization
        key = (campaign_raw, ad_group)
        campaign_name = campaign_name_map.get(key)
        if campaign_name is None:
            campaign_name = f"VitalFit_Google_Unknown_{ad_group}_Q1-2026"
            needs_review.append(("Google", campaign_raw, "Campaign_Name", f"Could not confidently map campaign '{campaign_raw}' / ad group '{ad_group}'."))

        ad_product = ad_type_map.get(ad_type)
        if ad_product is None:
            if ad_type == "PMAX":
                ad_product = "Standard_Display"
            else:
                ad_product = "NEEDS_REVIEW"
                needs_review.append(("Google", ad_type, "Ad_Product", f"Unknown Google ad type '{ad_type}'."))

        conv_type = conv_type_map.get(conv_type_raw)
        if conv_type is None:
            conv_type = "NEEDS_REVIEW"
            needs_review.append(("Google", conv_type_raw, "Conv_Type", f"Unknown conversion type '{conv_type_raw}'."))

        # Calculated metrics
        cpm = round((spend_media / impressions) * 1000, 2) if impressions > 0 else 0
        cpc = round(spend_media / clicks, 2) if clicks > 0 else 0
        cpa = round(spend_media / conversions, 2) if conversions > 0 else 0
        ctr = round((clicks / impressions) * 100, 2) if impressions > 0 else 0

        results.append({
            "Campaign_Name": campaign_name,
            "Platform": "Google",
            "Ad_Product": ad_product,
            "Date_Range": DATE_RANGE,
            "Impressions": impressions,
            "Clicks": clicks,
            "Spend_Media": spend_media,
            "Spend_Fees": spend_fees,
            "Spend_Total": spend_total,
            "Conversions": conversions,
            "Conv_Type": conv_type,
            "CPM": cpm,
            "CPC": cpc,
            "CPA": cpa,
            "CTR": ctr,
        })

    return results, mapping_decisions, needs_review


# ============================================================
# TRANSFORM TIKTOK DATA
# ============================================================
def transform_tiktok(rows):
    """
    TikTok: Total Cost = Spend_Total (INCLUDES platform fees).
    Fee structure is ambiguous -- not specified in the export or vendor reference.
    We flag this for human review and estimate platform fee at ~10% embedded in Total Cost.
    Result = Conversions. Optimization Event determines Conv_Type.
    """
    results = []
    mapping_decisions = []
    needs_review = []

    # TikTok fee assumption: platform takes ~10% of total cost as embedded fee
    # This is an estimate and should be verified with the vendor or contract
    ESTIMATED_PLATFORM_FEE_PCT = 0.10

    # Promotion type mapping
    promo_type_map = {
        "In-Feed Video": "Video",
        "TopView": "Video",
    }

    # Conv type mapping from Optimization Event
    conv_type_map = {
        "Complete Payment": "Purchase",
        "Add to Cart": "AddToCart",
    }

    # Campaign name mapping
    campaign_name_map = {
        ("VF_Spring_TT_Creators", "Creator Spark - Fitness"): "VitalFit_TikTok_Prospecting_CreatorFitness_Q1-2026",
        ("VF_Spring_TT_Creators", "Creator Spark - Wellness"): "VitalFit_TikTok_Prospecting_CreatorWellness_Q1-2026",
        ("VF_Spring_TT_Prospecting", "Broad Interest - Supplements"): "VitalFit_TikTok_Prospecting_BroadSupplements_Q1-2026",
        ("VF_Spring_TT_Prospecting", "Custom Audience - Gym"): "VitalFit_TikTok_Prospecting_CustomGym_Q1-2026",
        ("VF_Spring_TT_Retarget", "Site Visitors 7d"): "VitalFit_TikTok_Retargeting_SiteVisitors7d_Q1-2026",
        ("VF_Spring_TT_Retarget", "Video Viewers 50%"): "VitalFit_TikTok_Retargeting_VideoViewers50_Q1-2026",
    }

    # Record mapping decisions
    mapping_decisions.append(("TikTok", "Total Cost", "Spend_Total", "TikTok Total Cost includes platform fees per vendor reference."))
    mapping_decisions.append(("TikTok", "Platform Fee (est. 10%)", "Spend_Fees", "Fee structure NOT specified in export. Estimated at 10% of Total Cost. MUST be verified against TikTok contract or IO."))
    mapping_decisions.append(("TikTok", "Result", "Conversions", "Direct mapping per vendor reference."))
    mapping_decisions.append(("TikTok", "Optimization Event: Complete Payment", "Conv_Type: Purchase", "Mapped 'Complete Payment' to MDI 'Purchase'."))
    mapping_decisions.append(("TikTok", "Optimization Event: Add to Cart", "Conv_Type: AddToCart", "Mapped to MDI standard 'AddToCart'."))
    mapping_decisions.append(("TikTok", "Video Views", "(not mapped)", "No MDI equivalent per vendor reference unless counted as Impressions per SOW. Using Impressions column instead."))
    mapping_decisions.append(("TikTok", "Reach", "(not mapped)", "No MDI equivalent. Logged but excluded."))
    mapping_decisions.append(("TikTok", "CPC (vendor-provided)", "(recalculated)", "Vendor-provided CPC ignored. Recalculated from Spend_Media / Clicks for consistency."))
    mapping_decisions.append(("TikTok", "CPM (vendor-provided)", "(recalculated)", "Vendor-provided CPM ignored. Recalculated from Spend_Media for consistency."))
    mapping_decisions.append(("TikTok", "Promotion Type: In-Feed Video", "Ad_Product: Video", "In-Feed Video ads mapped to Video."))
    mapping_decisions.append(("TikTok", "Promotion Type: TopView", "Ad_Product: Video", "TopView is a premium video placement. Mapped to Video. Could warrant its own Ad_Product category."))

    # Flag the fee structure ambiguity
    needs_review.append(("TikTok", "(all campaigns)", "Spend_Fees", "TikTok fee structure is NOT documented in the raw export or vendor mapping reference. The 10% platform fee estimate is a placeholder. Verify against the TikTok Ads Manager contract, IO, or finance team before finalizing. This affects Spend_Media, CPM, CPC, and CPA for all TikTok rows."))
    needs_review.append(("TikTok", "TopView promotions", "Ad_Product", "TopView is a premium takeover placement that differs significantly from standard In-Feed Video. Consider whether MDI schema should have a separate category (e.g., 'Premium_Video' or 'Takeover')."))

    for row in rows:
        campaign_raw = row["Campaign"].strip()
        ad_group = row["Ad Group"].strip()
        promo_type = row["Promotion Type"].strip()
        impressions = safe_int(row["Impressions"])
        clicks = safe_int(row["Clicks"])
        total_cost = safe_float(row["Total Cost"])
        conversions = safe_int(row["Result"])
        opt_event = row["Optimization Event"].strip()

        # Fee estimation
        spend_total = total_cost
        spend_fees = round(total_cost * ESTIMATED_PLATFORM_FEE_PCT, 2)
        spend_media = round(spend_total - spend_fees, 2)

        # Campaign name standardization
        key = (campaign_raw, ad_group)
        campaign_name = campaign_name_map.get(key)
        if campaign_name is None:
            campaign_name = f"VitalFit_TikTok_Unknown_{ad_group}_Q1-2026"
            needs_review.append(("TikTok", campaign_raw, "Campaign_Name", f"Could not confidently map campaign '{campaign_raw}' / ad group '{ad_group}'."))

        ad_product = promo_type_map.get(promo_type)
        if ad_product is None:
            ad_product = "NEEDS_REVIEW"
            needs_review.append(("TikTok", promo_type, "Ad_Product", f"Unknown TikTok promotion type '{promo_type}'."))

        conv_type = conv_type_map.get(opt_event)
        if conv_type is None:
            conv_type = "NEEDS_REVIEW"
            needs_review.append(("TikTok", opt_event, "Conv_Type", f"Unknown optimization event '{opt_event}'."))

        # Calculated metrics from Spend_Media
        cpm = round((spend_media / impressions) * 1000, 2) if impressions > 0 else 0
        cpc = round(spend_media / clicks, 2) if clicks > 0 else 0
        cpa = round(spend_media / conversions, 2) if conversions > 0 else 0
        ctr = round((clicks / impressions) * 100, 2) if impressions > 0 else 0

        results.append({
            "Campaign_Name": campaign_name,
            "Platform": "TikTok",
            "Ad_Product": ad_product,
            "Date_Range": DATE_RANGE,
            "Impressions": impressions,
            "Clicks": clicks,
            "Spend_Media": spend_media,
            "Spend_Fees": spend_fees,
            "Spend_Total": spend_total,
            "Conversions": conversions,
            "Conv_Type": conv_type,
            "CPM": cpm,
            "CPC": cpc,
            "CPA": cpa,
            "CTR": ctr,
        })

    return results, mapping_decisions, needs_review


# ============================================================
# EXCEL FORMATTING HELPERS
# ============================================================
def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row_num, col_num, alt=False):
    cell = ws.cell(row=row_num, column=col_num)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if alt:
        cell.fill = ALT_ROW_FILL


def auto_fit_columns(ws, min_width=10, max_width=45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def apply_number_format(ws, col_num, start_row, end_row, fmt):
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=col_num).number_format = fmt


# ============================================================
# BUILD WORKBOOK
# ============================================================
def build_workbook():
    # Read raw data
    meta_rows = read_csv("raw-export-meta.csv")
    google_rows = read_csv("raw-export-google.csv")
    tiktok_rows = read_csv("raw-export-tiktok.csv")

    # Transform
    meta_data, meta_maps, meta_review = transform_meta(meta_rows)
    google_data, google_maps, google_review = transform_google(google_rows)
    tiktok_data, tiktok_maps, tiktok_review = transform_tiktok(tiktok_rows)

    all_data = meta_data + google_data + tiktok_data
    all_maps = meta_maps + google_maps + tiktok_maps
    all_review = meta_review + google_review + tiktok_review

    wb = openpyxl.Workbook()

    # --------------------------------------------------------
    # Sheet 1: MDI Standardized Data
    # --------------------------------------------------------
    ws1 = wb.active
    ws1.title = "MDI Standardized Data"

    # Title row
    ws1.merge_cells("A1:O1")
    title_cell = ws1.cell(row=1, column=1)
    title_cell.value = "MDI Standardized Data -- VitalFit Spring Campaign (Mar 1-15, 2026)"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Subtitle
    ws1.merge_cells("A2:O2")
    sub_cell = ws1.cell(row=2, column=1)
    sub_cell.value = "All vendor data (Meta, Google, TikTok) transformed to unified MDI format. See 'Mapping Decisions' and 'Needs Review' sheets for methodology."
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws1.row_dimensions[2].height = 22

    # Headers
    headers = [
        "Campaign_Name", "Platform", "Ad_Product", "Date_Range",
        "Impressions", "Clicks", "Spend_Media", "Spend_Fees", "Spend_Total",
        "Conversions", "Conv_Type", "CPM", "CPC", "CPA", "CTR"
    ]
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        ws1.cell(row=header_row, column=col_idx, value=header)
    style_header_row(ws1, header_row, len(headers))
    ws1.row_dimensions[header_row].height = 25

    # Data rows
    for row_idx, record in enumerate(all_data):
        excel_row = header_row + 1 + row_idx
        alt = row_idx % 2 == 1
        for col_idx, key in enumerate(headers, 1):
            val = record[key]
            ws1.cell(row=excel_row, column=col_idx, value=val)
            style_data_cell(ws1, excel_row, col_idx, alt=alt)

    data_start = header_row + 1
    data_end = header_row + len(all_data)

    # Number formatting
    # Impressions (col 5) and Clicks (col 6) - integer with comma separators
    apply_number_format(ws1, 5, data_start, data_end, "#,##0")
    apply_number_format(ws1, 6, data_start, data_end, "#,##0")
    # Spend columns (7, 8, 9) - currency
    for col in [7, 8, 9]:
        apply_number_format(ws1, col, data_start, data_end, '$#,##0.00')
    # Conversions (col 10) - integer
    apply_number_format(ws1, 10, data_start, data_end, "#,##0")
    # CPM (col 12), CPC (col 13), CPA (col 14) - currency
    for col in [12, 13, 14]:
        apply_number_format(ws1, col, data_start, data_end, '$#,##0.00')
    # CTR (col 15) - percentage
    # Convert CTR values to decimal for percentage format
    for row in range(data_start, data_end + 1):
        cell = ws1.cell(row=row, column=15)
        if cell.value is not None:
            cell.value = cell.value / 100  # Convert from 1.40 to 0.0140
        cell.number_format = '0.00%'

    # Freeze panes
    ws1.freeze_panes = "A5"

    # Auto-filter
    ws1.auto_filter.ref = f"A{header_row}:O{data_end}"

    auto_fit_columns(ws1, min_width=12, max_width=50)

    # --------------------------------------------------------
    # Sheet 2: Mapping Decisions
    # --------------------------------------------------------
    ws2 = wb.create_sheet("Mapping Decisions")

    ws2.merge_cells("A1:D1")
    title2 = ws2.cell(row=1, column=1)
    title2.value = "Mapping Decisions -- Vendor Term to MDI Standard"
    title2.font = TITLE_FONT
    title2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:D2")
    sub2 = ws2.cell(row=2, column=1)
    sub2.value = "Every translation from vendor-specific terminology to MDI standard fields, with rationale."
    sub2.font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws2.row_dimensions[2].height = 22

    map_headers = ["Vendor", "Vendor_Term", "MDI_Standard", "Decision / Rationale"]
    map_header_row = 4
    for col_idx, header in enumerate(map_headers, 1):
        ws2.cell(row=map_header_row, column=col_idx, value=header)
    style_header_row(ws2, map_header_row, len(map_headers))
    ws2.row_dimensions[map_header_row].height = 25

    for row_idx, (vendor, vendor_term, mdi_standard, rationale) in enumerate(all_maps):
        excel_row = map_header_row + 1 + row_idx
        alt = row_idx % 2 == 1
        ws2.cell(row=excel_row, column=1, value=vendor)
        ws2.cell(row=excel_row, column=2, value=vendor_term)
        ws2.cell(row=excel_row, column=3, value=mdi_standard)
        ws2.cell(row=excel_row, column=4, value=rationale)
        for col in range(1, 5):
            style_data_cell(ws2, excel_row, col, alt=alt)
            ws2.cell(row=excel_row, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    ws2.freeze_panes = "A5"
    auto_fit_columns(ws2, min_width=14, max_width=65)
    # Make rationale column wider
    ws2.column_dimensions["D"].width = 70

    # --------------------------------------------------------
    # Sheet 3: Needs Review
    # --------------------------------------------------------
    ws3 = wb.create_sheet("Needs Review")

    ws3.merge_cells("A1:D1")
    title3 = ws3.cell(row=1, column=1)
    title3.value = "Fields Requiring Human Review"
    title3.font = TITLE_FONT
    title3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 30

    ws3.merge_cells("A2:D2")
    sub3 = ws3.cell(row=2, column=1)
    sub3.value = "Items that could not be mapped with full confidence. Each requires verification before the MDI data is finalized."
    sub3.font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws3.row_dimensions[2].height = 22

    review_headers = ["Vendor", "Source Field/Campaign", "MDI Field", "Issue / Action Required"]
    review_header_row = 4
    for col_idx, header in enumerate(review_headers, 1):
        ws3.cell(row=review_header_row, column=col_idx, value=header)
    style_header_row(ws3, review_header_row, len(review_headers))
    ws3.row_dimensions[review_header_row].height = 25

    for row_idx, (vendor, source, mdi_field, issue) in enumerate(all_review):
        excel_row = review_header_row + 1 + row_idx
        ws3.cell(row=excel_row, column=1, value=vendor)
        ws3.cell(row=excel_row, column=2, value=source)
        ws3.cell(row=excel_row, column=3, value=mdi_field)
        ws3.cell(row=excel_row, column=4, value=issue)
        for col in range(1, 5):
            style_data_cell(ws3, excel_row, col)
            cell = ws3.cell(row=excel_row, column=col)
            cell.fill = REVIEW_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws3.freeze_panes = "A5"
    auto_fit_columns(ws3, min_width=14, max_width=65)
    ws3.column_dimensions["D"].width = 80

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------
    wb.save(OUTPUT_PATH)
    print(f"Output saved to: {OUTPUT_PATH}")
    print(f"  Sheet 1: MDI Standardized Data ({len(all_data)} rows)")
    print(f"  Sheet 2: Mapping Decisions ({len(all_maps)} entries)")
    print(f"  Sheet 3: Needs Review ({len(all_review)} items)")


if __name__ == "__main__":
    build_workbook()
