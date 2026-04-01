#!/usr/bin/env python3
"""
Billing & Invoice Reconciliation: Trace campaigns from Prisma Order through
CM360 placements to vendor invoices, identify discrepancies, and recommend actions.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "output.xlsx")

# ============================================================================
# Color palette and style constants
# ============================================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)

# Severity colors
HIGH_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
LOW_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

# Discrepancy highlight
DISCREPANCY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
MISSING_FILL = PatternFill(start_color="E6B8AF", end_color="E6B8AF", fill_type="solid")
OK_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CURRENCY_FORMAT = '$#,##0.00'
NUMBER_FORMAT = '#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, wrap=False):
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER


def auto_width(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                max_line = max(len(line) for line in lines)
                max_len = max(max_len, max_line)
        width = max(min_width, min(max_len + 3, max_width))
        ws.column_dimensions[col_letter].width = width


# ============================================================================
# SHEET 1: Chain of Custody
# ============================================================================
def build_chain_of_custody(wb):
    ws = wb.active
    ws.title = "Chain of Custody"

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "CHAIN OF CUSTODY -- VitalFit Spring Launch Campaign Reconciliation (March 2026)"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"].value = "Tracing each placement from Prisma Order through CM360 to Platform Invoice. Yellow = value diverges between systems. Red = missing from system."
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws.row_dimensions[2].height = 20

    # Section A: Meta Campaigns
    row = 4
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"].value = "SECTION A: META CAMPAIGNS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws[f"A{row}"].fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    row = 5
    headers = [
        "Canonical Name", "Prisma Line #", "CM360 Placement ID",
        "Prisma Order ($)", "CM360 Booked ($)", "Platform Actual Spend ($)",
        "Invoice Amount ($)", "Prisma vs CM360", "CM360 vs Invoice",
        "Notes / Flags"
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers))
    ws.row_dimensions[row].height = 35

    # Meta campaign data
    meta_data = [
        {
            "name": "VF_Prospecting_Broad_Feed",
            "prisma_line": "001",
            "cm360_id": "CM-44201",
            "prisma_amt": 28000.00,
            "cm360_amt": 28000.00,
            "platform_spend": 18800.00,
            "invoice_amt": 18800.00,
            "prisma_vs_cm360": "MATCH",
            "cm360_vs_invoice": "-$9,200 (partial month)",
            "notes": "Prisma flight starts Mar 10; invoice covers full March. Underspend expected -- campaign has Apr remaining."
        },
        {
            "name": "VF_Prospecting_LAL_Feed",
            "prisma_line": "002",
            "cm360_id": "CM-44202",
            "prisma_amt": 10500.00,
            "cm360_amt": 10500.00,
            "platform_spend": 7350.00,
            "invoice_amt": 7350.00,
            "prisma_vs_cm360": "MATCH",
            "cm360_vs_invoice": "-$3,150 (partial month)",
            "notes": "On pace for full flight. March = ~70% of order, consistent with 22/52 flight days."
        },
        {
            "name": "VF_Retargeting_ATC_Feed",
            "prisma_line": "003",
            "cm360_id": "CM-44203",
            "prisma_amt": 8400.00,
            "cm360_amt": 8400.00,
            "platform_spend": 6010.00,
            "invoice_amt": 6010.00,
            "prisma_vs_cm360": "MATCH",
            "cm360_vs_invoice": "-$2,390 (partial month)",
            "notes": "CM360 name drops '_Feed' suffix vs Prisma. Minor naming mismatch. Spend on track."
        },
        {
            "name": "VF_Retargeting_Purchaser_Stories",
            "prisma_line": "004",
            "cm360_id": "CM-44204",
            "prisma_amt": 3100.00,
            "cm360_amt": 3100.00,
            "platform_spend": 1920.00,
            "invoice_amt": 1920.00,
            "prisma_vs_cm360": "MATCH",
            "cm360_vs_invoice": "-$1,180 (partial month)",
            "notes": "Flight started Mar 17, only 15 days in March. Underpacing flagged. May need budget reallocation."
        },
        {
            "name": "VF_Prospecting_TikTok_Creators",
            "prisma_line": "N/A -- NOT IN PRISMA",
            "cm360_id": "CM-44205",
            "prisma_amt": None,
            "cm360_amt": 6000.00,
            "platform_spend": 0.00,
            "invoice_amt": 0.00,
            "prisma_vs_cm360": "MISSING IN PRISMA",
            "cm360_vs_invoice": "$0 (never activated)",
            "notes": "Marcus added Mar 14 per client request. Prisma update pending. No spend. Orphaned placement."
        },
    ]

    for i, d in enumerate(meta_data):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=d["name"])
        ws.cell(row=r, column=2, value=d["prisma_line"])
        ws.cell(row=r, column=3, value=d["cm360_id"])

        c4 = ws.cell(row=r, column=4)
        if d["prisma_amt"] is not None:
            c4.value = d["prisma_amt"]
            c4.number_format = CURRENCY_FORMAT
        else:
            c4.value = "N/A"
            c4.fill = MISSING_FILL

        c5 = ws.cell(row=r, column=5, value=d["cm360_amt"])
        c5.number_format = CURRENCY_FORMAT

        c6 = ws.cell(row=r, column=6, value=d["platform_spend"])
        c6.number_format = CURRENCY_FORMAT

        c7 = ws.cell(row=r, column=7, value=d["invoice_amt"])
        c7.number_format = CURRENCY_FORMAT

        c8 = ws.cell(row=r, column=8, value=d["prisma_vs_cm360"])
        if "MISSING" in d["prisma_vs_cm360"]:
            c8.fill = MISSING_FILL
        elif "MATCH" in d["prisma_vs_cm360"]:
            c8.fill = OK_FILL

        c9 = ws.cell(row=r, column=9, value=d["cm360_vs_invoice"])
        if d["cm360_amt"] and d["invoice_amt"] and d["cm360_amt"] != d["invoice_amt"]:
            c9.fill = DISCREPANCY_FILL

        ws.cell(row=r, column=10, value=d["notes"])

        for ci in range(1, 11):
            style_data_cell(ws.cell(row=r, column=ci), wrap=(ci == 10))

    # Totals for Meta
    total_row = row + 1 + len(meta_data)
    ws.cell(row=total_row, column=1, value="META SUBTOTAL").font = Font(name="Calibri", bold=True, size=10)
    for ci in [4, 5, 6, 7]:
        cell = ws.cell(row=total_row, column=ci)
        col_letter = get_column_letter(ci)
        cell.value = f"=SUM({col_letter}{row+1}:{col_letter}{total_row-1})"
        cell.number_format = CURRENCY_FORMAT
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.border = THIN_BORDER

    # Section B: Google Campaigns
    row = total_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"].value = "SECTION B: GOOGLE CAMPAIGNS (Direct Buy -- No CM360 Placements)"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws[f"A{row}"].fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    row += 1
    headers_google = [
        "Campaign Name", "Prisma Order ($)", "CM360 Booked ($)",
        "Platform Actual Spend ($)", "Invoice Amount ($)",
        "Order vs Invoice ($)", "Variance %", "Notes / Flags"
    ]
    for ci, h in enumerate(headers_google, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers_google))
    ws.row_dimensions[row].height = 35

    google_data = [
        {
            "name": "VF_Spring_Brand",
            "prisma_amt": 8000.00,
            "cm360_amt": None,
            "platform_spend": 7880.00,
            "invoice_amt": 7880.00,
            "variance_dollar": -120.00,
            "variance_pct": -1.5,
            "notes": "Direct buy, no CM360 placement. Slight underspend -- within tolerance."
        },
        {
            "name": "VF_Spring_NB_Protein",
            "prisma_amt": 9000.00,
            "cm360_amt": None,
            "platform_spend": 8820.00,
            "invoice_amt": 8820.00,
            "variance_dollar": -180.00,
            "variance_pct": -2.0,
            "notes": "Budget increased mid-month per internal record. Invoice below original order."
        },
        {
            "name": "VF_Spring_NB_Supplements",
            "prisma_amt": 9000.00,
            "cm360_amt": None,
            "platform_spend": 9120.00,
            "invoice_amt": 9120.00,
            "variance_dollar": 120.00,
            "variance_pct": 1.3,
            "notes": "Invoice EXCEEDS Prisma order by $120. Investigate: was budget increase authorized?"
        },
        {
            "name": "VF_Spring_Shopping",
            "prisma_amt": 8000.00,
            "cm360_amt": None,
            "platform_spend": 7930.00,
            "invoice_amt": 7930.00,
            "variance_dollar": -70.00,
            "variance_pct": -0.9,
            "notes": "Daily cap adjusted Mar 18. Underspend within acceptable range."
        },
        {
            "name": "VF_Spring_PMAX",
            "prisma_amt": 12000.00,
            "cm360_amt": None,
            "platform_spend": 12400.00,
            "invoice_amt": 12400.00,
            "variance_dollar": 400.00,
            "variance_pct": 3.3,
            "notes": "OVERSPEND: Budget increased Mar 15 per internal record. Prisma not updated. $400 over order."
        },
    ]

    for i, d in enumerate(google_data):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=d["name"])

        c2 = ws.cell(row=r, column=2, value=d["prisma_amt"])
        c2.number_format = CURRENCY_FORMAT

        c3 = ws.cell(row=r, column=3)
        c3.value = "N/A -- Direct Buy"
        c3.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        c4 = ws.cell(row=r, column=4, value=d["platform_spend"])
        c4.number_format = CURRENCY_FORMAT

        c5 = ws.cell(row=r, column=5, value=d["invoice_amt"])
        c5.number_format = CURRENCY_FORMAT

        c6 = ws.cell(row=r, column=6, value=d["variance_dollar"])
        c6.number_format = CURRENCY_FORMAT
        if d["variance_dollar"] > 0:
            c6.fill = DISCREPANCY_FILL

        c7 = ws.cell(row=r, column=7, value=f"{d['variance_pct']}%")
        if abs(d["variance_pct"]) > 2:
            c7.fill = DISCREPANCY_FILL

        ws.cell(row=r, column=8, value=d["notes"])

        for ci in range(1, 9):
            style_data_cell(ws.cell(row=r, column=ci), wrap=(ci == 8))

    # Google totals
    g_total_row = row + 1 + len(google_data)
    ws.cell(row=g_total_row, column=1, value="GOOGLE SUBTOTAL").font = Font(name="Calibri", bold=True, size=10)
    for ci in [2, 4, 5, 6]:
        cell = ws.cell(row=g_total_row, column=ci)
        col_letter = get_column_letter(ci)
        cell.value = f"=SUM({col_letter}{row+1}:{col_letter}{g_total_row-1})"
        cell.number_format = CURRENCY_FORMAT
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.border = THIN_BORDER

    # Grand total section
    gt_row = g_total_row + 2
    ws.merge_cells(f"A{gt_row}:J{gt_row}")
    ws[f"A{gt_row}"].value = "GRAND TOTALS"
    ws[f"A{gt_row}"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws[f"A{gt_row}"].fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    gt_row += 1
    summary_items = [
        ("Total Prisma Order (Meta)", 50000.00),
        ("Total Prisma Order (Google)", 46000.00),
        ("Combined Prisma Orders", 96000.00),
        ("Total March Invoiced (Meta)", 34080.00),
        ("Total March Invoiced (Google)", 46150.00),
        ("Combined March Invoiced", 80230.00),
        ("Invoiced vs Ordered (combined)", -15770.00),
        ("Agency Fees -- Meta (15%)", 5112.00),
        ("Agency Fees -- Google (flat $2,500/mo)", 2500.00),
        ("Total Client Charge (March)", 87842.00),
    ]
    for i, (label, val) in enumerate(summary_items):
        r = gt_row + i
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = CURRENCY_FORMAT
        c.font = Font(name="Calibri", size=10)
        c.border = THIN_BORDER
        ws.cell(row=r, column=1).border = THIN_BORDER

    auto_width(ws)
    ws.sheet_properties.tabColor = "1F4E79"


# ============================================================================
# SHEET 2: Discrepancy Analysis
# ============================================================================
def build_discrepancy_analysis(wb):
    ws = wb.create_sheet("Discrepancy Analysis")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "DISCREPANCY ANALYSIS -- All Identified Variances"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"].value = "Each row is a specific discrepancy between two systems. Sorted by severity."
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

    row = 4
    headers = [
        "#", "Placement / Campaign", "System A", "System B",
        "Dollar Difference", "Root Cause Hypothesis", "Severity",
        "Recommended Action"
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers))

    discrepancies = [
        {
            "num": 1,
            "placement": "VF_Prospecting_TikTok_Creators (CM-44205)",
            "system_a": "CM360: $6,000 booked",
            "system_b": "Prisma: NOT PRESENT",
            "diff": "$6,000 unreconciled",
            "root_cause": "Marcus created placement Mar 14 per client request. Prisma order was never updated. Placement was never activated on any platform, so $0 actual spend, but $6,000 sits in CM360 with no corresponding order authority.",
            "severity": "High",
            "action": "INVESTIGATE: Get Prisma order updated or remove CM360 placement. If client approved the TikTok spend, a Prisma amendment is needed before activation. If not approved, delete the CM360 placement to prevent accidental activation."
        },
        {
            "num": 2,
            "placement": "VF_Spring_PMAX (Google)",
            "system_a": "Prisma Order: $12,000",
            "system_b": "Invoice: $12,400",
            "diff": "+$400 overspend",
            "root_cause": "Budget was increased Mar 15 (per internal billing notes) but Prisma order was not amended. Google spent to the new budget, resulting in invoice exceeding authorized order amount.",
            "severity": "High",
            "action": "DISPUTE unless budget increase was formally approved. Get written client approval for the additional $400. Update Prisma order to reflect actual authorized budget. If approved, accept and amend order."
        },
        {
            "num": 3,
            "placement": "VF_Spring_NB_Supplements (Google)",
            "system_a": "Prisma Order: $9,000",
            "system_b": "Invoice: $9,120",
            "diff": "+$120 overspend",
            "root_cause": "No budget change noted in internal records, yet invoice exceeds order. Possible automated bidding overshoot or rounding on CPC billing. Small enough to be a platform pacing artifact.",
            "severity": "Medium",
            "action": "INVESTIGATE: Check Google Ads daily spend logs for the last week of March. If CPC bidding caused minor overshoot, accept and adjust daily caps. If no explanation, dispute the $120 difference."
        },
        {
            "num": 4,
            "placement": "Meta billing period vs Prisma flight dates",
            "system_a": "Meta Invoice: Mar 1-31",
            "system_b": "Prisma Flight: Mar 10 - Apr 30",
            "diff": "9 days of billing gap (Mar 1-9)",
            "root_cause": "Meta invoice covers full calendar month. Prisma flights start Mar 10. If Meta charged for any spend before Mar 10, those charges are unauthorized. Current data shows $0 pre-flight spend (campaigns likely not active before Mar 10), but the billing period mismatch is a process risk.",
            "severity": "Medium",
            "action": "ACCEPT current invoice (no evidence of pre-flight spend). Add a recurring check: compare platform activation dates vs order flight starts each billing cycle."
        },
        {
            "num": 5,
            "placement": "All Google campaigns -- no CM360 placements",
            "system_a": "CM360: 0 Google placements",
            "system_b": "Prisma/Invoice: 5 Google campaigns",
            "diff": "5 campaigns with no ad server tracking",
            "root_cause": "Google campaigns are managed as direct buys, bypassing CM360. This is a deliberate workflow choice but creates a gap in the chain of custody -- there is no third-party verification of delivery.",
            "severity": "Medium",
            "action": "ACCEPT for this cycle. Recommend adding Google campaigns to CM360 for unified reporting and third-party delivery verification. Flag as a process improvement for Q2."
        },
        {
            "num": 6,
            "placement": "VF_Retargeting_ATC naming mismatch",
            "system_a": "Prisma: VF_Retargeting_ATC_Feed",
            "system_b": "CM360: VF_Retargeting_ATC",
            "diff": "$0 (names differ, amounts match)",
            "root_cause": "CM360 placement name drops the '_Feed' suffix. Likely manual entry error when trafficking. Values match, so no financial impact, but inconsistent naming complicates automated reconciliation.",
            "severity": "Low",
            "action": "ACCEPT. Update CM360 placement name to match Prisma convention. Implement naming validation at trafficking step."
        },
        {
            "num": 7,
            "placement": "VF_Retargeting_Purchaser_Stories underpacing",
            "system_a": "Prisma Order: $3,100 (full flight)",
            "system_b": "March Invoice: $1,920 (62% of order)",
            "diff": "-$1,180 below order pace",
            "root_cause": "Flight started Mar 17 (7 days later than other lines). Only 15 active days in March vs 22 for other placements. Additionally flagged as underpacing in internal records. May need creative refresh or audience expansion.",
            "severity": "Low",
            "action": "ACCEPT. Monitor pacing in April. If still underpacing by Apr 15, recommend reallocating budget to higher-performing placements (Broad_Feed is overperforming)."
        },
        {
            "num": 8,
            "placement": "Agency fee calculation -- Meta vs Google",
            "system_a": "Meta: 15% of spend ($5,112)",
            "system_b": "Google: Flat $2,500/mo ($833.33/campaign)",
            "diff": "Different fee structures",
            "root_cause": "Internal billing record shows Meta agency fee at 15% of actual spend, while Google uses a flat $2,500/month split equally across 3 campaigns (internal record incorrectly splits across more). Inconsistent fee models can confuse client billing.",
            "severity": "Low",
            "action": "ACCEPT both fee structures (per contract). Verify Google flat fee is correctly allocated. Ensure client invoice clearly shows fee methodology for each vendor."
        },
    ]

    for i, d in enumerate(discrepancies):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=d["num"])
        ws.cell(row=r, column=2, value=d["placement"])
        ws.cell(row=r, column=3, value=d["system_a"])
        ws.cell(row=r, column=4, value=d["system_b"])
        ws.cell(row=r, column=5, value=d["diff"])
        ws.cell(row=r, column=6, value=d["root_cause"])
        ws.cell(row=r, column=7, value=d["severity"])
        ws.cell(row=r, column=8, value=d["action"])

        # Severity coloring for entire row
        severity_fill = {"High": HIGH_FILL, "Medium": MEDIUM_FILL, "Low": LOW_FILL}.get(d["severity"])
        for ci in range(1, 9):
            cell = ws.cell(row=r, column=ci)
            style_data_cell(cell, wrap=(ci in [5, 6, 8]))
            if severity_fill:
                cell.fill = severity_fill

        ws.row_dimensions[r].height = 80

    auto_width(ws, min_width=14, max_width=50)
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["H"].width = 55
    ws.sheet_properties.tabColor = "CC0000"


# ============================================================================
# SHEET 3: Name Mapping
# ============================================================================
def build_name_mapping(wb):
    ws = wb.create_sheet("Name Mapping")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "CAMPAIGN NAME MAPPING ACROSS SYSTEMS"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")

    ws.merge_cells("A2:F2")
    ws["A2"].value = "How the same campaign appears in Prisma, CM360, Meta/Google, and Internal Billing. Key for automated reconciliation."
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

    # Section A: Meta campaigns
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "META CAMPAIGNS"
    ws[f"A{row}"].font = SUBHEADER_FONT
    ws[f"A{row}"].fill = SUBHEADER_FILL

    row = 5
    headers = [
        "Canonical ID", "Prisma Name", "CM360 Name",
        "Meta Invoice Name", "Internal Billing Name", "Naming Notes"
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers))

    meta_names = [
        {
            "id": "BROAD_FEED",
            "prisma": "VF_Prospecting_Broad_Feed",
            "cm360": "VF_Prospecting_Broad_Feed",
            "meta": "VF_Spring_Prosp_Broad_18-35",
            "internal": "VF_Prospecting_Broad_Feed",
            "notes": "Meta adds 'Spring', abbreviates 'Prosp', appends age targeting '18-35'. Prisma/CM360/Internal match."
        },
        {
            "id": "LAL_FEED",
            "prisma": "VF_Prospecting_LAL_Feed",
            "cm360": "VF_Prospecting_LAL_Feed",
            "meta": "VF_Spring_Prosp_LAL",
            "internal": "VF_Prospecting_LAL_Feed",
            "notes": "Meta drops '_Feed' suffix and adds 'Spring'. Otherwise similar pattern."
        },
        {
            "id": "ATC_RETARGET",
            "prisma": "VF_Retargeting_ATC_Feed",
            "cm360": "VF_Retargeting_ATC",
            "meta": "VF_Spring_Retarget_ATC",
            "internal": "VF_Retargeting_ATC_Feed",
            "notes": "CM360 drops '_Feed'. Meta uses 'Retarget' instead of 'Retargeting' and adds 'Spring'. Three different name variants."
        },
        {
            "id": "PURCHASER_STORIES",
            "prisma": "VF_Retargeting_Purchaser_Stories",
            "cm360": "VF_Retargeting_Purchaser_Stories",
            "meta": "VF_Spring_Retarget_Purchaser",
            "internal": "VF_Retargeting_Purchaser_Stories",
            "notes": "Meta drops '_Stories' and uses 'Retarget'. Prisma/CM360/Internal match."
        },
        {
            "id": "TIKTOK_CREATORS",
            "prisma": "[NOT IN PRISMA]",
            "cm360": "VF_Prospecting_TikTok_Creators",
            "meta": "VF_Spring_Prosp_Creators",
            "internal": "VF_Prospecting_TikTok_Creators",
            "notes": "Orphaned: exists in CM360 and Meta (with $0 spend) but no Prisma order. Meta drops 'TikTok' from name."
        },
    ]

    for i, d in enumerate(meta_names):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=d["id"])
        ws.cell(row=r, column=2, value=d["prisma"])
        ws.cell(row=r, column=3, value=d["cm360"])
        ws.cell(row=r, column=4, value=d["meta"])
        ws.cell(row=r, column=5, value=d["internal"])
        ws.cell(row=r, column=6, value=d["notes"])

        for ci in range(1, 7):
            style_data_cell(ws.cell(row=r, column=ci), wrap=(ci == 6))

        # Highlight mismatches
        if d["prisma"] != d["cm360"] and "[NOT IN" not in d["prisma"]:
            ws.cell(row=r, column=3).fill = DISCREPANCY_FILL
        if "[NOT IN" in d["prisma"]:
            ws.cell(row=r, column=2).fill = MISSING_FILL

        ws.row_dimensions[r].height = 50

    # Section B: Google campaigns
    g_row = row + 1 + len(meta_names) + 1
    ws.merge_cells(f"A{g_row}:F{g_row}")
    ws[f"A{g_row}"].value = "GOOGLE CAMPAIGNS (Direct Buy -- No CM360)"
    ws[f"A{g_row}"].font = SUBHEADER_FONT
    ws[f"A{g_row}"].fill = SUBHEADER_FILL

    g_row += 1
    headers_g = [
        "Canonical ID", "Prisma Name", "CM360 Name",
        "Google Invoice Name", "Internal Billing Name", "Naming Notes"
    ]
    for ci, h in enumerate(headers_g, 1):
        ws.cell(row=g_row, column=ci, value=h)
    style_header_row(ws, g_row, len(headers_g))

    google_names = [
        {
            "id": "BRAND_SEARCH",
            "prisma": "VF_Spring_Brand",
            "cm360": "[NOT IN CM360]",
            "google": "VF_Spring_Brand",
            "internal": "VF_Spring_Brand",
            "notes": "Names match across Prisma, Google, and Internal. No CM360 entry (direct buy)."
        },
        {
            "id": "NB_PROTEIN",
            "prisma": "VF_Spring_NB_Protein",
            "cm360": "[NOT IN CM360]",
            "google": "VF_Spring_NB_Protein",
            "internal": "VF_Spring_NB_Protein",
            "notes": "Consistent naming. NB = Non-Brand."
        },
        {
            "id": "NB_SUPPLEMENTS",
            "prisma": "VF_Spring_NB_Supplements",
            "cm360": "[NOT IN CM360]",
            "google": "VF_Spring_NB_Supplements",
            "internal": "VF_Spring_NB_Supplements",
            "notes": "Consistent naming."
        },
        {
            "id": "SHOPPING",
            "prisma": "VF_Spring_Shopping",
            "cm360": "[NOT IN CM360]",
            "google": "VF_Spring_Shopping",
            "internal": "VF_Spring_Shopping",
            "notes": "Consistent naming."
        },
        {
            "id": "PMAX",
            "prisma": "VF_Spring_PMAX",
            "cm360": "[NOT IN CM360]",
            "google": "VF_Spring_PMAX",
            "internal": "VF_Spring_PMAX",
            "notes": "Consistent naming. PMAX = Performance Max."
        },
    ]

    for i, d in enumerate(google_names):
        r = g_row + 1 + i
        ws.cell(row=r, column=1, value=d["id"])
        ws.cell(row=r, column=2, value=d["prisma"])
        ws.cell(row=r, column=3, value=d["cm360"])
        ws.cell(row=r, column=4, value=d["google"])
        ws.cell(row=r, column=5, value=d["internal"])
        ws.cell(row=r, column=6, value=d["notes"])

        for ci in range(1, 7):
            style_data_cell(ws.cell(row=r, column=ci), wrap=(ci == 6))
        ws.cell(row=r, column=3).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws.row_dimensions[r].height = 40

    # Naming pattern summary
    p_row = g_row + 1 + len(google_names) + 2
    ws.merge_cells(f"A{p_row}:F{p_row}")
    ws[f"A{p_row}"].value = "NAMING PATTERN ANALYSIS"
    ws[f"A{p_row}"].font = SUBHEADER_FONT
    ws[f"A{p_row}"].fill = SUBHEADER_FILL

    patterns = [
        ("Meta naming convention", "VF_Spring_[Prosp/Retarget]_[Descriptor]_[Optional: targeting]. Adds 'Spring' campaign prefix, abbreviates objective, may append audience segment."),
        ("Prisma naming convention", "VF_[Objective]_[Descriptor]_[Format/Placement]. Uses full objective name (Prospecting/Retargeting), includes format hints (_Feed, _Stories)."),
        ("CM360 naming convention", "Generally follows Prisma but may drop suffixes (_Feed). Should match Prisma exactly -- deviations indicate manual entry errors."),
        ("Google naming convention", "VF_Spring_[Type]. Simpler structure. Matches Prisma because Google campaigns were set up from the order directly."),
        ("Reconciliation challenge", "Meta names share no common unique token with Prisma for 2 of 5 campaigns. Fuzzy matching on objective keywords (Prospecting/Prosp, Retargeting/Retarget, ATC, LAL, Purchaser, Broad) is required."),
        ("Recommendation", "Establish a shared campaign ID field across all systems. Use Prisma line number or a custom UTM-based ID that each platform preserves. Eliminates name-matching ambiguity entirely."),
    ]

    for i, (label, desc) in enumerate(patterns):
        r = p_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=desc).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 40
        for ci in range(1, 7):
            ws.cell(row=r, column=ci).border = THIN_BORDER

    auto_width(ws, min_width=16, max_width=45)
    ws.column_dimensions["F"].width = 60
    ws.sheet_properties.tabColor = "4A86C8"


# ============================================================================
# SHEET 4: Recommendations
# ============================================================================
def build_recommendations(wb):
    ws = wb.create_sheet("Recommendations")

    ws.merge_cells("A1:D1")
    ws["A1"].value = "RECOMMENDATIONS -- Actions and Process Improvements"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")

    ws.merge_cells("A2:D2")
    ws["A2"].value = "Prioritized list of immediate actions and long-term process fixes based on the reconciliation findings."
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

    # Section A: Immediate Actions
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = "IMMEDIATE ACTIONS (This Billing Cycle)"
    ws[f"A{row}"].font = SUBHEADER_FONT
    ws[f"A{row}"].fill = SUBHEADER_FILL

    row = 5
    headers = ["Priority", "Action Item", "Owner", "Details"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers))

    immediate_actions = [
        {
            "priority": "P1",
            "action": "Resolve CM-44205 (TikTok Creators) orphaned placement",
            "owner": "Marcus / Jordan",
            "details": "Either (a) get client written approval and create Prisma amendment for $6,000 TikTok line, or (b) delete the CM360 placement if not approved. Do NOT activate spend until Prisma order is updated. Deadline: before April billing cycle."
        },
        {
            "priority": "P1",
            "action": "Investigate PMAX overspend ($400 over Prisma order)",
            "owner": "Media Buyer / Jordan",
            "details": "Confirm whether the Mar 15 budget increase was client-approved. If yes: amend Prisma order to $12,400 and accept the invoice. If no: dispute $400 with Google and implement daily budget caps."
        },
        {
            "priority": "P2",
            "action": "Investigate NB_Supplements $120 overspend",
            "owner": "Media Buyer",
            "details": "Pull Google Ads daily spend report for March. Determine if CPC bidding caused overshoot. If platform pacing artifact: accept and tighten daily caps. If unexplained: dispute."
        },
        {
            "priority": "P2",
            "action": "Verify agency fee allocation for Google campaigns",
            "owner": "Billing Team",
            "details": "Internal billing splits $2,500 flat fee equally across 5 Google campaigns ($500 each). Original fee structure in internal record shows $833.33 per 3 campaigns. Clarify correct allocation methodology before sending client invoice."
        },
        {
            "priority": "P3",
            "action": "Accept Meta invoice ($34,080)",
            "owner": "Billing Team",
            "details": "All Meta line items match between CM360 booked amounts and invoice amounts for March spend. Differences from Prisma order are expected partial-month variances. Invoice is clean."
        },
        {
            "priority": "P3",
            "action": "Monitor VF_Retargeting_Purchaser_Stories pacing",
            "owner": "Media Buyer",
            "details": "At 62% of order value with only 15 active days in March, this line is underpacing. If still behind by Apr 15, recommend reallocating budget to Broad_Feed (which is pacing strong at 67% of order)."
        },
    ]

    for i, d in enumerate(immediate_actions):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=d["priority"])
        ws.cell(row=r, column=2, value=d["action"])
        ws.cell(row=r, column=3, value=d["owner"])
        ws.cell(row=r, column=4, value=d["details"])

        priority_fill = {"P1": HIGH_FILL, "P2": MEDIUM_FILL, "P3": LOW_FILL}.get(d["priority"])
        for ci in range(1, 5):
            cell = ws.cell(row=r, column=ci)
            style_data_cell(cell, wrap=(ci == 4))
            if priority_fill:
                cell.fill = priority_fill
        ws.row_dimensions[r].height = 70

    # Section B: Process Improvements
    pi_row = row + 1 + len(immediate_actions) + 1
    ws.merge_cells(f"A{pi_row}:D{pi_row}")
    ws[f"A{pi_row}"].value = "PROCESS IMPROVEMENTS (Q2 and Beyond)"
    ws[f"A{pi_row}"].font = SUBHEADER_FONT
    ws[f"A{pi_row}"].fill = SUBHEADER_FILL

    pi_row += 1
    headers_pi = ["#", "Improvement", "Impact", "Implementation Notes"]
    for ci, h in enumerate(headers_pi, 1):
        ws.cell(row=pi_row, column=ci, value=h)
    style_header_row(ws, pi_row, len(headers_pi))

    improvements = [
        {
            "num": 1,
            "improvement": "Implement shared Campaign ID across all systems",
            "impact": "Eliminates name-matching ambiguity. Currently requires fuzzy matching across 3 different naming conventions.",
            "notes": "Use Prisma line number or a custom ID (e.g., VF-S26-001) as a mandatory field in CM360, Meta Ads Manager, Google Ads, and internal billing. Add as UTM parameter for attribution tracking."
        },
        {
            "num": 2,
            "improvement": "Add Google campaigns to CM360",
            "impact": "Creates complete chain of custody for all vendors. Currently Google has no third-party delivery verification.",
            "notes": "Set up CM360 Floodlight tags for Google campaigns. Even if managed as direct buy, CM360 tracking provides independent impression/click verification. Estimated setup: 2 hours per campaign."
        },
        {
            "num": 3,
            "improvement": "Enforce Prisma-first workflow for all new placements",
            "impact": "Prevents orphaned placements like CM-44205. No CM360 placement should be created without a corresponding Prisma line.",
            "notes": "Add a workflow gate: CM360 placement creation requires Prisma line number. If client requests an urgent addition, create a provisional Prisma line immediately (even if $0 budget) and amend later."
        },
        {
            "num": 4,
            "improvement": "Standardize naming convention across platforms",
            "impact": "Reduces reconciliation time and errors. Currently Meta uses a completely different naming pattern.",
            "notes": "Proposed format: [Client]_[Season][Year]_[Objective]_[Descriptor]_[Platform]. Example: VF_S26_Prosp_Broad_Meta. Apply at campaign creation, enforce via naming template in each platform."
        },
        {
            "num": 5,
            "improvement": "Automate monthly reconciliation report",
            "impact": "Catches discrepancies within days instead of weeks. Current process is manual and error-prone.",
            "notes": "Build a script that ingests Prisma export, CM360 export, and vendor invoices. Auto-matches on Campaign ID (see improvement #1), flags variances above threshold (e.g., 2% or $500). Output: this spreadsheet, generated automatically."
        },
        {
            "num": 6,
            "improvement": "Align billing periods with flight dates",
            "impact": "Prevents confusion from partial-month comparisons. March invoice vs Mar 10-Apr 30 flight makes variance analysis misleading.",
            "notes": "Request platform-level flight-date billing where available. For Meta: not possible (calendar month billing). Workaround: prorate Prisma order amounts to the billing month for accurate comparison."
        },
        {
            "num": 7,
            "improvement": "Clarify and document agency fee structures",
            "impact": "Prevents billing disputes and internal confusion about fee calculations.",
            "notes": "Create a fee schedule document for each client. For VitalFit: Meta = 15% of spend, Google = flat $2,500/month. Define how flat fees are allocated across campaigns. Review annually."
        },
    ]

    for i, d in enumerate(improvements):
        r = pi_row + 1 + i
        ws.cell(row=r, column=1, value=d["num"])
        ws.cell(row=r, column=2, value=d["improvement"])
        ws.cell(row=r, column=3, value=d["impact"])
        ws.cell(row=r, column=4, value=d["notes"])

        for ci in range(1, 5):
            style_data_cell(ws.cell(row=r, column=ci), wrap=(ci in [2, 3, 4]))
        ws.row_dimensions[r].height = 70

    # Dispute vs Accept Summary
    s_row = pi_row + 1 + len(improvements) + 2
    ws.merge_cells(f"A{s_row}:D{s_row}")
    ws[f"A{s_row}"].value = "DISPUTE vs ACCEPT SUMMARY"
    ws[f"A{s_row}"].font = SUBHEADER_FONT
    ws[f"A{s_row}"].fill = SUBHEADER_FILL

    s_row += 1
    headers_s = ["Vendor", "Line Item", "Amount", "Decision"]
    for ci, h in enumerate(headers_s, 1):
        ws.cell(row=s_row, column=ci, value=h)
    style_header_row(ws, s_row, len(headers_s))

    decisions = [
        ("Meta", "VF_Spring_Prosp_Broad_18-35", 18800.00, "ACCEPT -- matches platform spend"),
        ("Meta", "VF_Spring_Prosp_LAL", 7350.00, "ACCEPT -- matches platform spend"),
        ("Meta", "VF_Spring_Retarget_ATC", 6010.00, "ACCEPT -- matches platform spend"),
        ("Meta", "VF_Spring_Retarget_Purchaser", 1920.00, "ACCEPT -- matches, monitor pacing"),
        ("Meta", "VF_Spring_Prosp_Creators", 0.00, "ACCEPT -- $0, never activated"),
        ("Google", "VF_Spring_Brand", 7880.00, "ACCEPT -- within tolerance"),
        ("Google", "VF_Spring_NB_Protein", 8820.00, "ACCEPT -- within tolerance"),
        ("Google", "VF_Spring_NB_Supplements", 9120.00, "INVESTIGATE -- $120 over order"),
        ("Google", "VF_Spring_Shopping", 7930.00, "ACCEPT -- within tolerance"),
        ("Google", "VF_Spring_PMAX", 12400.00, "DISPUTE -- $400 over order, pending approval"),
    ]

    for i, (vendor, item, amt, decision) in enumerate(decisions):
        r = s_row + 1 + i
        ws.cell(row=r, column=1, value=vendor)
        ws.cell(row=r, column=2, value=item)
        c3 = ws.cell(row=r, column=3, value=amt)
        c3.number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=4, value=decision)

        for ci in range(1, 5):
            style_data_cell(ws.cell(row=r, column=ci), wrap=(ci == 4))

        if "DISPUTE" in decision:
            for ci in range(1, 5):
                ws.cell(row=r, column=ci).fill = HIGH_FILL
        elif "INVESTIGATE" in decision:
            for ci in range(1, 5):
                ws.cell(row=r, column=ci).fill = MEDIUM_FILL

    auto_width(ws, min_width=14, max_width=45)
    ws.column_dimensions["D"].width = 60
    ws.sheet_properties.tabColor = "38761D"


# ============================================================================
# Main
# ============================================================================
def main():
    wb = openpyxl.Workbook()
    build_chain_of_custody(wb)
    build_discrepancy_analysis(wb)
    build_name_mapping(wb)
    build_recommendations(wb)
    wb.save(OUTPUT_PATH)
    print(f"Output saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
