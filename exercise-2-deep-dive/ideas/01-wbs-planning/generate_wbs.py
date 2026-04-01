import sys
sys.path.insert(0, '/Users/hammermt/.claude/plugins/marketplaces/anthropic-agent-skills/skills/xlsx')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Colors ---
header_fill = PatternFill('solid', fgColor='1F4E79')
phase_fill = PatternFill('solid', fgColor='2E75B6')
flag_fill = PatternFill('solid', fgColor='FFF2CC')
header_font = Font(bold=True, color='FFFFFF', size=11)
phase_font = Font(bold=True, color='FFFFFF', size=12)
flag_font = Font(bold=True, color='BF8F00', size=10)
normal_font = Font(size=10)
bold_font = Font(bold=True, size=10)
title_font = Font(bold=True, size=16, color='1F4E79')
subtitle_font = Font(size=11, color='404040')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
wrap = Alignment(wrap_text=True, vertical='top')

# ===== SHEET 1: WBS =====
ws = wb.active
ws.title = 'WBS'
ws.sheet_properties.tabColor = '1F4E79'

ws.merge_cells('A1:E1')
ws['A1'] = 'VitalFit Supplements — Spring Launch WBS'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:E2')
ws['A2'] = 'Client: VitalFit  |  Budget: $400K  |  Duration: 10 weeks  |  Channels: Meta, TikTok, Google Search, Connected TV'
ws['A2'].font = subtitle_font
ws.row_dimensions[2].height = 20

cols = {'A': 40, 'B': 28, 'C': 14, 'D': 30, 'E': 70}
for c, w in cols.items():
    ws.column_dimensions[c].width = w

headers = ['Task', 'Owner', 'Duration', 'Dependencies', 'Notes']

phases = [
    ('Phase 1: Setup (Weeks 1-3)', [
        ['Client kickoff call', 'Sarah (Account Lead)', 'Week 1', '—', 'First engagement — set expectations on CTV measurement, creative timelines, and legal review process'],
        ['Receive brand assets & guidelines', 'Sarah (Account Lead)', 'Week 1', 'Kickoff call', '[FLAG] Creative is from external agency we haven\'t worked with — confirm asset specs and handoff process at kickoff'],
        ['Audience research & persona development', 'Anika (Insights)', 'Week 1-2', 'Kickoff call', 'New client — essential for targeting across 4 channels. Inform TikTok creator strategy'],
        ['[FLAG] Legal review of supplement ad claims', 'Sarah (Account Lead) + Client Legal', 'Week 1-2', 'Kickoff call', 'NEW STEP — no precedent in past projects. Supplement claims must be cleared before any copy goes live. Start early to avoid blocking launch'],
        ['Build media plan', 'Jordan (Planner)', 'Week 1-3', 'Kickoff call; Audience research', '4 channels = 3-week build (matches GlowUp pattern). Heavy TikTok allocation (50% / ~$200K) needs dedicated planning'],
        ['[FLAG] TikTok creator sourcing & briefing', 'Jordan (Planner) + External Creative', 'Week 2-3', 'Media plan (draft); Legal review', 'NEW STEP. 50% of budget is creator-driven TikTok — need creator selection, briefing, and content calendar. No past project had this scale of creator work'],
        ['[FLAG] CTV inventory & deal setup', 'Marcus (Programmatic Lead)', 'Week 2-3', 'Media plan (draft)', 'NEW CHANNEL. Agency hasn\'t run CTV for DTC before. Marcus to source inventory, negotiate CPMs, and define targeting. Build in extra buffer'],
        ['[FLAG] CTV measurement & attribution setup', 'Marcus (Programmatic Lead) + Anika (Insights)', 'Week 2-3', 'CTV deal setup', 'NEW STEP. Client nervous about CTV measurement — define KPIs, attribution methodology, and reporting approach before launch'],
        ['Trafficking & tagging (all platforms)', 'Priya (Ad Ops)', 'Week 3', 'Media plan; Legal review; Brand assets', '4 platforms — full week needed (matches GlowUp). Legal-approved copy required before trafficking'],
    ]),
    ('Phase 2: Launch (Weeks 4-5)', [
        ['Phased launch: Meta + Google Search', 'Jordan (Planner) + Priya (Ad Ops)', 'Week 4', 'Trafficking & tagging', 'Lead with proven channels first (matches GlowUp staggered approach)'],
        ['TikTok launch (creator content)', 'Jordan (Planner) + Priya (Ad Ops)', 'Week 5', 'Phased launch; Creator content ready', 'Stagger by 1 week to QA creator assets and manage pacing on largest budget line'],
        ['[FLAG] CTV launch', 'Marcus (Programmatic Lead) + Priya (Ad Ops)', 'Week 5', 'CTV deal setup; Trafficking', 'Launch alongside TikTok. Treat as test — start with conservative pacing and scale based on early signals'],
    ]),
    ('Phase 3: Optimize (Weeks 4-9)', [
        ['Weekly performance reviews', 'Jordan (Planner)', 'Weeks 4-9', 'Campaign launch', 'Include CTV metrics even if early data is thin — client will want visibility'],
        ['[FLAG] CTV early-read report', 'Marcus (Programmatic Lead) + Anika (Insights)', 'Week 6', '2 weeks of CTV data', 'NEW STEP. Client is nervous about CTV — provide a dedicated early read with go/no-go recommendation before mid-flight'],
        ['Mid-flight optimization report', 'Jordan (Planner) + Anika (Insights)', 'Week 7', 'Weekly reviews', 'Full cross-channel view. Assess CAC vs. targets ($18 Meta, $22 TikTok). Recommend budget shifts if needed'],
        ['Creative refresh (Meta + TikTok)', 'External Creative + Jordan (Planner)', 'Week 7', 'Mid-flight report', '[FLAG] External creative agency — build in extra lead time. Request refreshed assets by Week 6 to launch Week 7'],
        ['Audience & budget rebalancing', 'Jordan (Planner)', 'Week 7-8', 'Mid-flight report', 'Expect potential shift between TikTok and CTV based on performance. GlowUp shifted Display to TikTok at this stage'],
    ]),
    ('Phase 4: Wrap (Weeks 9-11)', [
        ['Final performance report', 'Jordan (Planner) + Anika (Insights)', 'Week 10', 'All campaigns complete', 'Include CTV learnings section — this will be the agency\'s first DTC CTV case study'],
        ['Vendor reconciliation', 'Finance + Marcus (Programmatic Lead)', 'Week 10-11', 'Final report', '4 vendors + new CTV partner = higher reconciliation complexity. FreshBrew took an extra week on Meta alone — budget 2 weeks'],
        ['Client debrief + next steps', 'Sarah (Account Lead) + Jordan (Planner)', 'Week 11', 'Final report; Vendor reconciliation', 'Include CTV test results and recommendation for future investment. New client — use debrief to build relationship for renewal'],
    ]),
]

row = 4
for phase_name, tasks in phases:
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = phase_name
    ws[f'A{row}'].font = phase_font
    ws[f'A{row}'].fill = phase_fill
    ws[f'A{row}'].alignment = Alignment(vertical='center')
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = phase_fill
    ws.row_dimensions[row].height = 28
    row += 1

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[row].height = 22
    row += 1

    for task in tasks:
        is_flag = '[FLAG]' in task[0] or '[FLAG]' in task[4]
        for ci, val in enumerate(task, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = flag_font if is_flag and ci == 1 else normal_font
            cell.fill = flag_fill if is_flag else PatternFill()
            cell.alignment = wrap
            cell.border = thin_border
        ws.row_dimensions[row].height = 45
        row += 1
    row += 1

# Uncertainty Flags
row += 1
ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'] = 'Uncertainty Flags'
ws[f'A{row}'].font = phase_font
ws[f'A{row}'].fill = PatternFill('solid', fgColor='C00000')
for c in range(1, 6):
    ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor='C00000')
ws.row_dimensions[row].height = 28
row += 1

for ci, h in enumerate(['Flag', 'Risk', 'Mitigation'], 1):
    cell = ws.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
row += 1

flags = [
    ['CTV is brand new for the agency (DTC context)', 'No internal benchmarks for pacing, CPMs, or attribution. Marcus is capable but flying without a playbook', 'Start conservative, build a dedicated CTV early-read checkpoint at Week 6, and document everything for future projects'],
    ['External creative agency (unknown partner)', 'Asset delays, format mismatches, or misaligned creative quality — FreshBrew had a 2-day delay with an internal creative partner', 'Establish asset specs + delivery SLA at kickoff. Request creative refresh assets 1 week before needed. Sarah to own the relationship'],
    ['Legal review for supplement claims', 'Entirely new compliance step — could block ad copy and delay trafficking if not started early', 'Begin legal review in Week 1 in parallel with audience research. Don\'t let it become a Week 3 bottleneck'],
    ['TikTok at 50% of budget ($200K) — creator-driven', 'Past TikTok work (GlowUp) was not creator-heavy. Creator sourcing, briefing, and content approval is a different workflow', 'Dedicate planning time in Weeks 2-3. Jordan should own a creator content calendar separate from the media plan'],
    ['Aggressive CAC targets ($18 Meta / $22 TikTok)', 'No historical data from this client. DTC supplements is a competitive category', 'Set expectations at kickoff that Weeks 4-5 are a learning phase. Mid-flight report should include a realistic CAC trajectory, not just a snapshot'],
    ['CTV measurement / attribution', 'Client is already nervous. If reporting is unclear, trust erodes on a channel that\'s eating budget', 'Define measurement methodology before launch (Week 2-3). Anika and Marcus co-own this. Present it to the client before go-live'],
]

for flag in flags:
    ws.merge_cells(f'B{row}:C{row}')
    ws.merge_cells(f'D{row}:E{row}')
    ws.cell(row=row, column=1, value=flag[0]).font = bold_font
    ws.cell(row=row, column=1).alignment = wrap
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=flag[1]).alignment = wrap
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=4, value=flag[2]).alignment = wrap
    ws.cell(row=row, column=4).border = thin_border
    ws.row_dimensions[row].height = 55
    row += 1

# ===== SHEET 2: GANTT CHART =====
gs = wb.create_sheet('Gantt Chart')
gs.sheet_properties.tabColor = '2E75B6'

# Gantt data: (task, owner, phase, start_week, end_week, is_flag)
gantt_tasks = [
    ('Client kickoff call', 'Sarah', 'Setup', 1, 1, False),
    ('Receive brand assets & guidelines', 'Sarah', 'Setup', 1, 1, True),
    ('Audience research & persona development', 'Anika', 'Setup', 1, 2, False),
    ('Legal review of supplement ad claims', 'Sarah + Legal', 'Setup', 1, 2, True),
    ('Build media plan', 'Jordan', 'Setup', 1, 3, False),
    ('TikTok creator sourcing & briefing', 'Jordan + Creative', 'Setup', 2, 3, True),
    ('CTV inventory & deal setup', 'Marcus', 'Setup', 2, 3, True),
    ('CTV measurement & attribution setup', 'Marcus + Anika', 'Setup', 2, 3, True),
    ('Trafficking & tagging (all platforms)', 'Priya', 'Setup', 3, 3, False),
    ('Phased launch: Meta + Google Search', 'Jordan + Priya', 'Launch', 4, 4, False),
    ('TikTok launch (creator content)', 'Jordan + Priya', 'Launch', 5, 5, False),
    ('CTV launch', 'Marcus + Priya', 'Launch', 5, 5, True),
    ('Weekly performance reviews', 'Jordan', 'Optimize', 4, 9, False),
    ('CTV early-read report', 'Marcus + Anika', 'Optimize', 6, 6, True),
    ('Mid-flight optimization report', 'Jordan + Anika', 'Optimize', 7, 7, False),
    ('Creative refresh (Meta + TikTok)', 'External + Jordan', 'Optimize', 7, 7, True),
    ('Audience & budget rebalancing', 'Jordan', 'Optimize', 7, 8, False),
    ('Final performance report', 'Jordan + Anika', 'Wrap', 10, 10, False),
    ('Vendor reconciliation', 'Finance + Marcus', 'Wrap', 10, 11, False),
    ('Client debrief + next steps', 'Sarah + Jordan', 'Wrap', 11, 11, False),
]

phase_colors = {
    'Setup': 'BDD7EE',
    'Launch': 'A9D18E',
    'Optimize': 'FFD966',
    'Wrap': 'F4B183',
}
phase_bar_colors = {
    'Setup': '2E75B6',
    'Launch': '548235',
    'Optimize': 'BF8F00',
    'Wrap': 'C55A11',
}

gs.column_dimensions['A'].width = 42
gs.column_dimensions['B'].width = 20
gs.column_dimensions['C'].width = 12
for w in range(1, 12):
    gs.column_dimensions[get_column_letter(w + 3)].width = 12

gs.merge_cells('A1:N1')
gs['A1'] = 'VitalFit Supplements — Gantt Chart (10 Weeks + Wrap)'
gs['A1'].font = title_font
gs['A1'].alignment = Alignment(vertical='center')
gs.row_dimensions[1].height = 30

gs.merge_cells('A2:N2')
gs['A2'] = 'Budget: $400K  |  Channels: Meta, TikTok, Google Search, Connected TV  |  Yellow rows = deviation flags'
gs['A2'].font = subtitle_font
gs.row_dimensions[2].height = 20

# Header row
row = 4
gantt_headers = ['Task', 'Owner', 'Phase'] + [f'Week {w}' for w in range(1, 12)]
for ci, h in enumerate(gantt_headers, 1):
    cell = gs.cell(row=row, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
gs.row_dimensions[row].height = 24

# Phase separator tracking
current_phase = None
row = 5

for task_name, owner, phase, start, end, is_flag in gantt_tasks:
    if phase != current_phase:
        current_phase = phase
        gs.merge_cells(f'A{row}:N{row}')
        phase_label = {
            'Setup': 'Phase 1: Setup (Weeks 1-3)',
            'Launch': 'Phase 2: Launch (Weeks 4-5)',
            'Optimize': 'Phase 3: Optimize (Weeks 4-9)',
            'Wrap': 'Phase 4: Wrap (Weeks 9-11)',
        }[phase]
        gs[f'A{row}'] = phase_label
        gs[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=11)
        pf = PatternFill('solid', fgColor=phase_bar_colors[phase])
        for c in range(1, 15):
            gs.cell(row=row, column=c).fill = pf
        gs.row_dimensions[row].height = 22
        row += 1

    row_fill = flag_fill if is_flag else PatternFill('solid', fgColor=phase_colors[phase])
    task_font = flag_font if is_flag else normal_font

    gs.cell(row=row, column=1, value=task_name).font = task_font
    gs.cell(row=row, column=1).fill = row_fill
    gs.cell(row=row, column=1).alignment = wrap
    gs.cell(row=row, column=1).border = thin_border

    gs.cell(row=row, column=2, value=owner).font = normal_font
    gs.cell(row=row, column=2).fill = row_fill
    gs.cell(row=row, column=2).alignment = wrap
    gs.cell(row=row, column=2).border = thin_border

    gs.cell(row=row, column=3, value=phase).font = normal_font
    gs.cell(row=row, column=3).fill = row_fill
    gs.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    gs.cell(row=row, column=3).border = thin_border

    for w in range(1, 12):
        col = w + 3
        cell = gs.cell(row=row, column=col)
        cell.border = thin_border
        if start <= w <= end:
            cell.fill = PatternFill('solid', fgColor=phase_bar_colors[phase])
            if w == start and start == end:
                cell.value = '●'
            elif w == start:
                cell.value = '◄'
            elif w == end:
                cell.value = '►'
            else:
                cell.value = '■'
            cell.font = Font(color='FFFFFF', bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.fill = row_fill

    gs.row_dimensions[row].height = 26
    row += 1

# Legend
row += 2
gs[f'A{row}'] = 'Legend'
gs[f'A{row}'].font = Font(bold=True, size=12, color='1F4E79')
row += 1
legend_items = [
    ('Setup', phase_bar_colors['Setup']),
    ('Launch', phase_bar_colors['Launch']),
    ('Optimize', phase_bar_colors['Optimize']),
    ('Wrap', phase_bar_colors['Wrap']),
]
for label, color in legend_items:
    gs.cell(row=row, column=1, value=label).font = bold_font
    gs.cell(row=row, column=2).fill = PatternFill('solid', fgColor=color)
    gs.cell(row=row, column=2).border = thin_border
    row += 1

gs.cell(row=row, column=1, value='Deviation Flag').font = flag_font
gs.cell(row=row, column=1).fill = flag_fill
gs.cell(row=row, column=1).border = thin_border

# Freeze panes
gs.freeze_panes = 'D5'
ws.freeze_panes = 'A4'

# Print settings
for sheet in [ws, gs]:
    sheet.sheet_format.defaultRowHeight = 15
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1

output_path = '/Users/hammermt/Codes/zenith-aof/exercise-2-deep-dive/ideas/01-wbs-planning/vitalfit-wbs.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
