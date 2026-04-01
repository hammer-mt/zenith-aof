"""
Budget QA Discrepancy Report Generator
Compares master budget sheet against Google Ads and Meta platform exports,
cross-references change log for context, and outputs a formatted Excel report.
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_PATH = os.path.join(BASE_DIR, "output.xlsx")

# ------------------------------------------------------------------
# 1. Load data
# ------------------------------------------------------------------

def load_csv(filename):
    with open(os.path.join(DATA_DIR, filename), newline="") as f:
        return list(csv.DictReader(f))

master = load_csv("master-budget-sheet.csv")
google = load_csv("google-ads-budget-export.csv")
meta = load_csv("meta-budget-export.csv")

# Index platform data by campaign name
google_by_campaign = {r["Campaign"]: r for r in google}
meta_by_campaign = {}
for r in meta:
    # Meta export has "VF_Spring_Prosp_Broad_18-35" but master has "VF_Spring_Prosp_Broad"
    name = r["Campaign"]
    meta_by_campaign[name] = r

# Manual mapping: Meta export uses a different campaign name for Prosp_Broad
# "VF_Spring_Prosp_Broad_18-35" maps to master "VF_Spring_Prosp_Broad"
META_NAME_MAP = {
    "VF_Spring_Prosp_Broad": "VF_Spring_Prosp_Broad_18-35",
}

# ------------------------------------------------------------------
# 2. Change log evidence (pre-parsed)
# ------------------------------------------------------------------

change_log = {
    "VF_Spring_NB_Protein": {
        "date": "Mar 12",
        "who": "Jordan (Planner)",
        "note": "Client approved on call. Jordan said he would update master sheet tomorrow -- master not yet updated.",
        "has_client_approval": True,
        "master_updated": False,
    },
    "VF_Spring_PMAX": {
        "date": "Mar 15",
        "who": "Jordan (Planner)",
        "note": "PMAX crushing it, client said go for it. Increased to $433/day.",
        "has_client_approval": True,
        "master_updated": False,
    },
    "VF_Spring_Shopping": {
        "date": "Mar 18",
        "who": "Priya (Ad Ops)",
        "note": "Adjusted daily cap up to $300 to help underspending. No client approval mentioned.",
        "has_client_approval": False,
        "master_updated": False,
    },
    "VF_Spring_Prosp_Broad": {
        "date": "Mar 14",
        "who": "Jordan (Planner)",
        "note": "Increased by $33/day. Sarah has the approval email from client.",
        "has_client_approval": True,
        "master_updated": False,
    },
    "VF_Spring_Retarget_ATC": {
        "date": "Mar 10",
        "who": "Priya (Ad Ops)",
        "note": "Bumped by $20/day. Waiting on Jordan to confirm with client.",
        "has_client_approval": False,
        "master_updated": False,
    },
    "LS_Q1_Awareness": {
        "date": "Mar 05",
        "who": "Sarah (Account Lead)",
        "note": "Client asked to pull back awareness spend -- drop $500/month. Meta updated.",
        "has_client_approval": True,
        "master_updated": False,
    },
    "LS_Q1_Search_NB": {
        "date": "Mar 20",
        "who": "Sarah (Account Lead)",
        "note": "Client wants to pause for last 10 days of Q1, shifting budget to Meta awareness.",
        "has_client_approval": True,
        "master_updated": False,
    },
}


# ------------------------------------------------------------------
# 3. Build discrepancy report rows
# ------------------------------------------------------------------

def get_platform_channel(channel_str):
    """Return 'Google' or 'Meta' based on channel string."""
    if "Google" in channel_str or "PMAX" in channel_str.upper():
        return "Google"
    elif "Meta" in channel_str:
        return "Meta"
    return "Unknown"


def find_platform_record(campaign_name, channel):
    """Find the matching platform record for a campaign."""
    platform = get_platform_channel(channel)
    if platform == "Google":
        return google_by_campaign.get(campaign_name), platform
    elif platform == "Meta":
        # Try direct match first, then mapped name
        mapped_name = META_NAME_MAP.get(campaign_name, campaign_name)
        rec = meta_by_campaign.get(campaign_name) or meta_by_campaign.get(mapped_name)
        return rec, platform
    return None, platform


discrepancy_rows = []
missing_rows = []
matched_platform_campaigns_google = set()
matched_platform_campaigns_meta = set()

for m in master:
    client = m["Client"]
    campaign = m["Campaign"]
    channel = m["Channel"]
    master_budget = float(m["Monthly_Budget"])
    platform = get_platform_channel(channel)

    plat_rec, _ = find_platform_record(campaign, channel)

    if plat_rec is None:
        # Campaign in master but not in platform -- check if this is expected
        # (e.g., Meta campaigns won't be in Google export and vice versa)
        # This is expected behavior, not an issue
        status = "Match (Expected)"
        classification = "N/A"
        evidence = "Campaign is on {} -- not expected in other platform exports.".format(channel)
        priority = "N/A"
        discrepancy_rows.append({
            "client": client,
            "campaign": campaign,
            "channel": channel,
            "master_budget": master_budget,
            "platform_budget": master_budget,  # assume match
            "diff_dollar": 0,
            "diff_pct": 0,
            "status": "Match",
            "platform_status": "N/A (not in this platform export)",
            "classification": "N/A",
            "evidence": "No platform export mismatch -- budget assumed correct at master level.",
            "priority": "N/A",
        })
        # But we DO want to flag if a campaign that SHOULD be in a platform export is missing
        # Actually, let's just check: is the campaign in the correct platform export?
        # We already tried to find it. If not found, it's missing from its own platform.
        # Let's redo this logic more carefully.
        missing_rows.append({
            "client": client,
            "campaign": campaign,
            "channel": channel,
            "master_budget": master_budget,
            "direction": "In Master, Missing from {} Export".format(platform),
            "notes": "Campaign exists in master sheet but was not found in the {} budget export pulled Mar 26.".format(platform),
        })
        # Remove the last discrepancy row we just added -- we'll handle missing separately
        discrepancy_rows.pop()
        continue

    # Track matched campaigns
    if platform == "Google":
        matched_platform_campaigns_google.add(plat_rec["Campaign"])
    else:
        matched_platform_campaigns_meta.add(plat_rec["Campaign"])

    platform_budget = float(plat_rec["Monthly_Equivalent"])
    platform_status = plat_rec.get("Status", "Unknown")
    diff_dollar = platform_budget - master_budget
    diff_pct = (diff_dollar / master_budget * 100) if master_budget != 0 else 0

    # Determine status
    if abs(diff_dollar) < 0.01 and platform_status == "Active":
        status = "Match"
    elif platform_status == "Paused":
        status = "Mismatch (Paused)"
    elif abs(diff_dollar) >= 0.01:
        status = "Mismatch"
    else:
        status = "Match"

    # Classification and evidence from change log
    cl = change_log.get(campaign)
    if status == "Match":
        classification = "OK"
        evidence = "Budgets align."
        priority = "None"
    elif cl:
        if cl["has_client_approval"]:
            classification = "Likely Approved"
            evidence = "{}: {} -- {}".format(cl["date"], cl["who"], cl["note"])
            priority = "Low" if abs(diff_dollar) <= 500 else "Medium"
        else:
            classification = "Needs Investigation"
            evidence = "{}: {} -- {}".format(cl["date"], cl["who"], cl["note"])
            priority = "High" if abs(diff_dollar) >= 500 else "Medium"
    else:
        classification = "Error"
        evidence = "No change log entry found for this discrepancy."
        priority = "High"

    # Special handling for paused campaigns
    if platform_status == "Paused" and cl and cl["has_client_approval"]:
        classification = "Likely Approved"
        priority = "Medium"

    discrepancy_rows.append({
        "client": client,
        "campaign": campaign,
        "channel": channel,
        "master_budget": master_budget,
        "platform_budget": platform_budget,
        "diff_dollar": diff_dollar,
        "diff_pct": round(diff_pct, 1),
        "status": status,
        "platform_status": platform_status,
        "classification": classification,
        "evidence": evidence,
        "priority": priority,
    })

# Check for campaigns in platform exports but NOT in master
# (reverse direction -- platform has it, master doesn't)
for g_campaign, g_rec in google_by_campaign.items():
    if g_campaign not in matched_platform_campaigns_google:
        # Not matched -- check if it's truly missing from master
        master_names = [m["Campaign"] for m in master]
        if g_campaign not in master_names:
            missing_rows.append({
                "client": "Unknown",
                "campaign": g_campaign,
                "channel": "Google",
                "master_budget": "N/A",
                "direction": "In Google Export, Missing from Master",
                "notes": "Campaign found in Google Ads export but not in master budget sheet.",
            })

for m_campaign, m_rec in meta_by_campaign.items():
    if m_campaign not in matched_platform_campaigns_meta:
        # Check reverse mapping too
        reverse_mapped = False
        for master_name, mapped_name in META_NAME_MAP.items():
            if mapped_name == m_campaign:
                reverse_mapped = True
                break
        if not reverse_mapped:
            master_names = [m["Campaign"] for m in master]
            if m_campaign not in master_names:
                missing_rows.append({
                    "client": "Unknown",
                    "campaign": m_campaign,
                    "channel": "Meta",
                    "master_budget": "N/A",
                    "direction": "In Meta Export, Missing from Master",
                    "notes": "Campaign found in Meta export but not in master budget sheet.",
                })


# ------------------------------------------------------------------
# 4. Build summary data
# ------------------------------------------------------------------

total_exposure = sum(abs(r["diff_dollar"]) for r in discrepancy_rows if r["status"] != "Match")
count_match = sum(1 for r in discrepancy_rows if r["status"] == "Match")
count_mismatch = sum(1 for r in discrepancy_rows if "Mismatch" in r["status"])
count_missing = len(missing_rows)

count_high = sum(1 for r in discrepancy_rows if r["priority"] == "High")
count_medium = sum(1 for r in discrepancy_rows if r["priority"] == "Medium")
count_low = sum(1 for r in discrepancy_rows if r["priority"] == "Low")

# Build action items
actions = []
for r in discrepancy_rows:
    if r["status"] == "Match":
        continue
    if r["classification"] == "Likely Approved":
        actions.append({
            "campaign": r["campaign"],
            "action": "Update master sheet to reflect approved change (${:,.0f} -> ${:,.0f}). Verify approval documentation is on file.".format(
                r["master_budget"], r["platform_budget"]
            ),
            "owner": "Account Lead / Planner",
            "priority": r["priority"],
            "classification": r["classification"],
        })
    elif r["classification"] == "Needs Investigation":
        actions.append({
            "campaign": r["campaign"],
            "action": "Confirm client approval for budget change (${:,.0f} -> ${:,.0f}). If unapproved, revert platform budget immediately.".format(
                r["master_budget"], r["platform_budget"]
            ),
            "owner": "Planner (Jordan)",
            "priority": r["priority"],
            "classification": r["classification"],
        })
    elif r["classification"] == "Error":
        actions.append({
            "campaign": r["campaign"],
            "action": "Investigate unexplained discrepancy. No change log entry. Revert platform budget to master (${:,.0f}) or obtain retroactive approval.".format(
                r["master_budget"]
            ),
            "owner": "Ad Ops + Account Lead",
            "priority": r["priority"],
            "classification": r["classification"],
        })

for m in missing_rows:
    actions.append({
        "campaign": m["campaign"],
        "action": "Resolve missing campaign: {}. Verify whether campaign should exist in this system.".format(m["direction"]),
        "owner": "Ad Ops",
        "priority": "Medium",
        "classification": "Missing Campaign",
    })


# ------------------------------------------------------------------
# 5. Create Excel workbook
# ------------------------------------------------------------------

wb = Workbook()

# -- Styles --
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
CURRENCY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.0"%"'

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="404040")
METRIC_FONT = Font(name="Calibri", bold=True, size=12)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_body_cell(ws, row, col, value=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = BODY_FONT
    cell.alignment = BODY_ALIGN
    cell.border = THIN_BORDER
    return cell


def auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                lengths.append(max(len(l) for l in lines))
        if lengths:
            w = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = w


# ============================================================
# Sheet 1: Discrepancy Report
# ============================================================

ws1 = wb.active
ws1.title = "Discrepancy Report"

# Title
ws1.merge_cells("A1:L1")
title_cell = ws1["A1"]
title_cell.value = "Budget QA -- Discrepancy Report"
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:L2")
sub_cell = ws1["A2"]
sub_cell.value = "Generated: April 1, 2026 | Data pulled: March 26, 2026 | Master sheet last updated: March 25, 2026"
sub_cell.font = SUBTITLE_FONT

HEADERS_1 = [
    "Client", "Campaign", "Channel", "Master Budget", "Platform Budget",
    "Difference ($)", "Difference (%)", "Platform Status", "Status",
    "Classification", "Evidence", "Priority"
]

header_row = 4
for col_idx, h in enumerate(HEADERS_1, 1):
    ws1.cell(row=header_row, column=col_idx, value=h)
style_header_row(ws1, header_row, len(HEADERS_1))

for i, r in enumerate(discrepancy_rows):
    row = header_row + 1 + i
    style_body_cell(ws1, row, 1, r["client"])
    style_body_cell(ws1, row, 2, r["campaign"])
    style_body_cell(ws1, row, 3, r["channel"])

    c4 = style_body_cell(ws1, row, 4, r["master_budget"])
    c4.number_format = CURRENCY_FORMAT

    c5 = style_body_cell(ws1, row, 5, r["platform_budget"])
    c5.number_format = CURRENCY_FORMAT

    c6 = style_body_cell(ws1, row, 6, r["diff_dollar"])
    c6.number_format = '+#,##0.00;-#,##0.00;0.00'

    c7 = style_body_cell(ws1, row, 7, r["diff_pct"])
    c7.number_format = '+0.0"%";-0.0"%";"0.0%"'

    style_body_cell(ws1, row, 8, r["platform_status"])
    style_body_cell(ws1, row, 9, r["status"])
    style_body_cell(ws1, row, 10, r["classification"])
    style_body_cell(ws1, row, 11, r["evidence"])
    style_body_cell(ws1, row, 12, r["priority"])

    # Conditional fill based on status/classification
    if r["status"] == "Match":
        fill = GREEN_FILL
    elif r["classification"] == "Likely Approved":
        fill = YELLOW_FILL
    elif r["classification"] in ("Needs Investigation", "Error"):
        fill = RED_FILL
    else:
        fill = None

    if fill:
        for col in range(1, len(HEADERS_1) + 1):
            ws1.cell(row=row, column=col).fill = fill

# Freeze top rows
ws1.freeze_panes = "A5"
ws1.auto_filter.ref = "A4:L{}".format(header_row + len(discrepancy_rows))
auto_width(ws1)

# ============================================================
# Sheet 2: Missing Campaigns
# ============================================================

ws2 = wb.create_sheet("Missing Campaigns")

ws2.merge_cells("A1:F1")
ws2["A1"].value = "Missing Campaigns -- Cross-Platform Reconciliation"
ws2["A1"].font = TITLE_FONT
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:F2")
ws2["A2"].value = "Campaigns present in one system but not found in the corresponding platform export."
ws2["A2"].font = SUBTITLE_FONT

HEADERS_2 = ["Client", "Campaign", "Channel", "Master Budget", "Direction", "Notes"]
h_row2 = 4
for col_idx, h in enumerate(HEADERS_2, 1):
    ws2.cell(row=h_row2, column=col_idx, value=h)
style_header_row(ws2, h_row2, len(HEADERS_2))

if missing_rows:
    for i, r in enumerate(missing_rows):
        row = h_row2 + 1 + i
        style_body_cell(ws2, row, 1, r["client"])
        style_body_cell(ws2, row, 2, r["campaign"])
        style_body_cell(ws2, row, 3, r["channel"])
        budget_val = r["master_budget"]
        c = style_body_cell(ws2, row, 4, budget_val if budget_val == "N/A" else float(budget_val))
        if budget_val != "N/A":
            c.number_format = CURRENCY_FORMAT
        style_body_cell(ws2, row, 5, r["direction"])
        style_body_cell(ws2, row, 6, r["notes"])

        for col in range(1, len(HEADERS_2) + 1):
            ws2.cell(row=row, column=col).fill = YELLOW_FILL
else:
    row = h_row2 + 1
    ws2.merge_cells("A{}:F{}".format(row, row))
    ws2["A{}".format(row)].value = "No missing campaigns detected. All master campaigns have corresponding platform records."
    ws2["A{}".format(row)].font = Font(name="Calibri", size=10, italic=True)

ws2.freeze_panes = "A5"
auto_width(ws2)

# ============================================================
# Sheet 3: Summary & Actions
# ============================================================

ws3 = wb.create_sheet("Summary & Actions")

ws3.merge_cells("A1:F1")
ws3["A1"].value = "Summary & Recommended Actions"
ws3["A1"].font = TITLE_FONT
ws3.row_dimensions[1].height = 30

# Key Metrics section
r = 3
ws3.cell(row=r, column=1, value="KEY METRICS").font = Font(name="Calibri", bold=True, size=12, color="2F5496")
ws3.merge_cells("A{}:B{}".format(r, r))

metrics = [
    ("Total Campaigns Audited", len(discrepancy_rows) + len(missing_rows)),
    ("Campaigns Matching", count_match),
    ("Campaigns with Discrepancies", count_mismatch),
    ("Missing from Platform Exports", count_missing),
    ("Total Financial Exposure", "${:,.2f}".format(total_exposure)),
    ("High Priority Issues", count_high),
    ("Medium Priority Issues", count_medium),
    ("Low Priority Issues", count_low),
]

for i, (label, val) in enumerate(metrics):
    row = r + 1 + i
    c1 = ws3.cell(row=row, column=1, value=label)
    c1.font = Font(name="Calibri", size=10, bold=True)
    c1.border = THIN_BORDER
    c1.fill = LIGHT_GRAY_FILL
    c2 = ws3.cell(row=row, column=2, value=val)
    c2.font = METRIC_FONT
    c2.border = THIN_BORDER
    c2.alignment = Alignment(horizontal="center")

    # Color code severity counts
    if label == "High Priority Issues" and val > 0:
        c2.fill = RED_FILL
    elif label == "Medium Priority Issues" and val > 0:
        c2.fill = YELLOW_FILL
    elif label == "Campaigns Matching":
        c2.fill = GREEN_FILL

# Discrepancy Breakdown
breakdown_start = r + 1 + len(metrics) + 2
ws3.cell(row=breakdown_start, column=1, value="DISCREPANCY BREAKDOWN").font = Font(
    name="Calibri", bold=True, size=12, color="2F5496"
)
ws3.merge_cells("A{}:F{}".format(breakdown_start, breakdown_start))

bd_headers = ["Campaign", "Channel", "Master", "Platform", "Difference", "Classification"]
bh_row = breakdown_start + 1
for col_idx, h in enumerate(bd_headers, 1):
    ws3.cell(row=bh_row, column=col_idx, value=h)
style_header_row(ws3, bh_row, len(bd_headers))

mismatches = [r for r in discrepancy_rows if r["status"] != "Match"]
for i, r in enumerate(mismatches):
    row = bh_row + 1 + i
    style_body_cell(ws3, row, 1, r["campaign"])
    style_body_cell(ws3, row, 2, r["channel"])
    c3 = style_body_cell(ws3, row, 3, r["master_budget"])
    c3.number_format = CURRENCY_FORMAT
    c4 = style_body_cell(ws3, row, 4, r["platform_budget"])
    c4.number_format = CURRENCY_FORMAT
    c5 = style_body_cell(ws3, row, 5, r["diff_dollar"])
    c5.number_format = '+#,##0.00;-#,##0.00;0.00'
    style_body_cell(ws3, row, 6, r["classification"])

    if r["classification"] == "Likely Approved":
        fill = YELLOW_FILL
    elif r["classification"] in ("Needs Investigation", "Error"):
        fill = RED_FILL
    else:
        fill = None
    if fill:
        for col in range(1, len(bd_headers) + 1):
            ws3.cell(row=row, column=col).fill = fill

# Recommended Actions
action_start = bh_row + 1 + len(mismatches) + 2
ws3.cell(row=action_start, column=1, value="RECOMMENDED ACTIONS").font = Font(
    name="Calibri", bold=True, size=12, color="2F5496"
)
ws3.merge_cells("A{}:F{}".format(action_start, action_start))

act_headers = ["Priority", "Campaign", "Classification", "Action Required", "Owner"]
ah_row = action_start + 1
for col_idx, h in enumerate(act_headers, 1):
    ws3.cell(row=ah_row, column=col_idx, value=h)
style_header_row(ws3, ah_row, len(act_headers))

# Sort actions: High first, then Medium, then Low
priority_order = {"High": 0, "Medium": 1, "Low": 2}
actions_sorted = sorted(actions, key=lambda x: priority_order.get(x["priority"], 3))

for i, a in enumerate(actions_sorted):
    row = ah_row + 1 + i
    style_body_cell(ws3, row, 1, a["priority"])
    style_body_cell(ws3, row, 2, a["campaign"])
    style_body_cell(ws3, row, 3, a["classification"])
    style_body_cell(ws3, row, 4, a["action"])
    style_body_cell(ws3, row, 5, a["owner"])

    if a["priority"] == "High":
        fill = RED_FILL
    elif a["priority"] == "Medium":
        fill = YELLOW_FILL
    else:
        fill = GREEN_FILL
    for col in range(1, len(act_headers) + 1):
        ws3.cell(row=row, column=col).fill = fill

# Process note at bottom
note_row = ah_row + 1 + len(actions_sorted) + 2
ws3.merge_cells("A{}:F{}".format(note_row, note_row))
ws3["A{}".format(note_row)].value = (
    "PROCESS NOTE: All budget changes should follow the approval workflow: "
    "Client approval -> Master sheet update -> Platform adjustment. "
    "Discrepancies where platform was changed without updating the master sheet "
    "indicate a process gap even when the change itself was approved."
)
ws3["A{}".format(note_row)].font = Font(name="Calibri", size=9, italic=True, color="666666")
ws3["A{}".format(note_row)].alignment = Alignment(wrap_text=True)

ws3.freeze_panes = "A3"
auto_width(ws3, max_width=55)
ws3.column_dimensions["D"].width = 60  # Action column needs more space

# ------------------------------------------------------------------
# 6. Save
# ------------------------------------------------------------------

wb.save(OUTPUT_PATH)
print("Report saved to: {}".format(OUTPUT_PATH))
print("  Sheet 1: Discrepancy Report ({} campaigns)".format(len(discrepancy_rows)))
print("  Sheet 2: Missing Campaigns ({} items)".format(len(missing_rows)))
print("  Sheet 3: Summary & Actions ({} action items)".format(len(actions_sorted)))
print("  Total financial exposure: ${:,.2f}".format(total_exposure))
